"""⓪ 입력 검증 (설계서 v0.3 §4.1) — 파이프라인 진입 전 실행.

사용:
  python validate_inputs.py --checklist <xlsx> [--root <폴더> | --from-listing <json>]
  python validate_inputs.py --root <폴더>            # 모드 B (폴더분만)

결과를 JSON으로 출력한다: {"findings":[{code, action, detail}], "target_sheets":[...]}
exit code: FATAL 존재 시 2, PAUSE 존재 시 1, 그 외 0(WARN 포함).
동작 정의의 정본은 references/output_contract.md 오류 코드 분류표.
"""

from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl

from config import DRAFT_META_SHEET, FATAL, PAUSE, SHEET_PREFIX, WARN


@dataclass(frozen=True, slots=True)
class Finding:
    code: str
    action: str  # FATAL | PAUSE | WARN
    detail: str


def normalize_sheet_name(name: str) -> str:
    """대소문자·전각·앞뒤 공백 무시 정규화."""
    return unicodedata.normalize("NFKC", name).strip().lower()


def check_checklist(path: Path) -> tuple[list[Finding], list[str]]:
    findings: list[Finding] = []
    target_sheets: list[str] = []
    if not path.is_file():
        return [Finding("CHECKLIST_UNREADABLE", FATAL, f"파일 없음: {path}")], []
    try:
        workbook = openpyxl.load_workbook(path, read_only=True)
    except Exception as error:  # noqa: BLE001 — 열기 실패 원인 그대로 보고
        return [Finding("CHECKLIST_UNREADABLE", FATAL, f"열기 실패: {error}")], []

    prefix = normalize_sheet_name(SHEET_PREFIX)
    target_sheets = [
        name for name in workbook.sheetnames
        if normalize_sheet_name(name).startswith(prefix)
    ]
    if DRAFT_META_SHEET in workbook.sheetnames:
        meta = {}
        try:
            rows = list(workbook[DRAFT_META_SHEET].iter_rows(values_only=True))
            meta = {str(r[0]): r[1] for r in rows if r and r[0] is not None}
        except Exception:  # noqa: BLE001
            pass
        if str(meta.get("reviewed", "false")).lower() != "true":
            findings.append(Finding(
                "AI_DRAFT_CHECKLIST", PAUSE,
                f"모드 B 초안 마커 검출(미검수 상태: {meta.get('generated_at', '?')} 생성) — "
                "사람 검수 완료를 확인해야 진행 가능",
            ))
        else:
            findings.append(Finding(
                "AI_DRAFT_CHECKLIST", WARN,
                "검수 완료된 초안 출신 체크리스트 — source.draft_origin에 이력 기록 필요",
            ))
    if not target_sheets:
        findings.append(Finding(
            "TARGET_SHEET_NOT_FOUND", PAUSE,
            f"Target* 시트 0개 — 모드 B(폴더 추론) 제안. 시트 목록: {workbook.sheetnames}",
        ))
    elif len(target_sheets) > 1:
        findings.append(Finding(
            "MULTIPLE_TARGET_SHEETS", PAUSE,
            f"Target* 시트 {len(target_sheets)}개 — 사용자 선택 필요: {target_sheets}",
        ))
    return findings, target_sheets


CHECKLIST_NAME_HINTS = ("체크리스트", "checklist")


def find_checklist_candidates(root: Path) -> list[str]:
    """모드 B 폴더 안의 체크리스트 후보(.xlsx에 Target* 시트 보유)를 찾는다 (v0.4 §4.0)."""
    prefix = normalize_sheet_name(SHEET_PREFIX)
    candidates: list[str] = []
    for path in sorted(root.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        except Exception:  # noqa: BLE001 — 못 여는 파일은 후보 아님
            continue
        if any(normalize_sheet_name(name).startswith(prefix) for name in sheet_names):
            candidates.append(str(path.relative_to(root)))
        if len(candidates) >= 3:  # 안내 목적 — 상한
            break
    return candidates


def find_checklist_candidates_by_name(listing: Path) -> list[str]:
    """listing 모드 — 파일을 열 수 없어 이름 힌트로만 후보를 찾는다 (약한 휴리스틱)."""
    entries = json.loads(listing.read_text(encoding="utf-8")).get("entries", [])
    return [
        str(e.get("name")) for e in entries
        if e.get("type") == "file" and str(e.get("name", "")).lower().endswith(".xlsx")
        and any(hint in str(e.get("name", "")).lower() for hint in CHECKLIST_NAME_HINTS)
    ][:3]


def check_folder(root: Path | None, listing: Path | None,
                 mode_b: bool = False) -> list[Finding]:
    findings: list[Finding] = []
    if listing is not None:
        if not listing.is_file():
            return [Finding("DELIVERABLE_ROOT_NOT_FOUND", FATAL, f"listing 파일 없음: {listing}")]
        entries = json.loads(listing.read_text(encoding="utf-8")).get("entries", [])
        if not any(e.get("type") == "file" for e in entries):
            return [Finding("DELIVERABLE_ROOT_EMPTY", PAUSE, "listing에 파일 0건 — 경로 재확인 필요")]
        if mode_b:
            candidates = find_checklist_candidates_by_name(listing)
            if candidates:
                findings.append(Finding(
                    "CHECKLIST_FOUND_IN_FOLDER", PAUSE,
                    f"폴더에 체크리스트로 보이는 파일 발견(이름 기준 추정): {candidates} — "
                    "이 파일로 모드 A 진행 여부를 사용자에게 확인 (자동 전환 금지)",
                ))
        return findings
    if root is None:
        return []
    if not root.is_dir():
        return [Finding("DELIVERABLE_ROOT_NOT_FOUND", FATAL, f"폴더 없음: {root}")]
    if not any(p.is_file() for p in root.rglob("*")):
        return [Finding("DELIVERABLE_ROOT_EMPTY", PAUSE, "폴더에 파일 0건 — 경로 재확인 필요 (전 항목 MISSING 양산 금지)")]
    if mode_b:
        candidates = find_checklist_candidates(root)
        if candidates:
            findings.append(Finding(
                "CHECKLIST_FOUND_IN_FOLDER", PAUSE,
                f"폴더 안에서 Target* 시트 보유 .xlsx 발견: {candidates} — "
                "이 파일로 모드 A 진행 여부를 사용자에게 확인 (자동 전환 금지)",
            ))
    return findings


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="⓪ 입력 검증")
    parser.add_argument("--checklist", default=None)
    parser.add_argument("--root", default=None)
    parser.add_argument("--from-listing", default=None)
    args = parser.parse_args(argv)

    findings: list[Finding] = []
    target_sheets: list[str] = []
    if args.checklist:
        checklist_findings, target_sheets = check_checklist(Path(args.checklist))
        findings.extend(checklist_findings)
    findings.extend(check_folder(
        Path(args.root) if args.root else None,
        Path(args.from_listing) if args.from_listing else None,
        mode_b=args.checklist is None,  # v0.4: 모드 B 사전 검증 시 체크리스트 후보 탐지
    ))

    print(json.dumps(
        {"findings": [asdict(f) for f in findings], "target_sheets": target_sheets},
        ensure_ascii=False, indent=2,
    ))
    if any(f.action == FATAL for f in findings):
        return 2
    if any(f.action == PAUSE for f in findings):
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
