"""모드 B B9 — 확정 mapping.json에서 Target 시트 초안 .xlsx 역생성.

사용: python generate_target_draft.py mapping.json -o Target시트초안.xlsx
- gate confirmed + mode=folder_inferred 에서만 생성
- 동일 절차그룹+동일 표준 산출물의 복수 파일은 한 행에 개행 다중 기재
- AI 초안 마커: 숨김 시트 _ai_draft_meta (reviewed=false)
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from config import DRAFT_META_SHEET, MODE_FOLDER, SKILL_NAME, SKILL_VERSION

logger = logging.getLogger(__name__)

HEADERS = ["No.", "프로세스/절차/지침/가이드", "출력 작업 산출물",
           "파일 명 또는 관련 URL", "비고", "Note"]


def build_rows(mapping: dict[str, Any]) -> list[list[str]]:
    """(그룹, 산출물) 단위로 파일을 묶는다 — 사람 확정값(human_decision) 우선."""
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for item in mapping["items"]:
        inference = item["inference"]
        group = inference.get("human_decision") or item.get("procedure_group") or "(미분류 — 검수 필요)"
        product = item.get("work_product") or "(산출물명 검수 필요)"
        key = (group, product)
        bucket = buckets.setdefault(key, {"files": [], "bases": []})
        path = item["evidence"][0]["resolved_path"]
        bucket["files"].append(path)
        if inference.get("basis"):
            bucket["bases"].append(inference["basis"])
    rows: list[list[str]] = []
    for number, ((group, product), bucket) in enumerate(sorted(buckets.items()), start=1):
        basis = "; ".join(sorted(set(bucket["bases"]))[:2]) or "키워드 사전 미매칭 — 사람 분류 필요"
        rows.append([
            str(number), group, product, "\n".join(bucket["files"]),
            f"[AI 추론 초안 — 확정: 점검자] 근거: {basis}", "",
        ])
    return rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Target 시트 초안 역생성 (모드 B)")
    parser.add_argument("mapping")
    parser.add_argument("--phase", default=None, help="시트명에 쓸 단계명 (기본: project.phase)")
    parser.add_argument("-o", "--output", required=True)
    args = parser.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")

    mapping = json.loads(Path(args.mapping).read_text(encoding="utf-8"))
    if mapping["source"].get("mode") != MODE_FOLDER:
        raise SystemExit("[거부] 모드 B(folder_inferred) mapping에서만 초안을 생성합니다")
    if mapping["gate"].get("status") != "confirmed":
        raise SystemExit("[거부] 사용자 확정(confirmed) 후에만 초안을 생성합니다 — 분류 검수 선행")

    phase = args.phase or mapping.get("project", {}).get("phase", "")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Target_{phase}"[:31] if phase else "Target_초안"
    sheet["A1"] = "점검 대상 (AI 추론 초안 — 검수 후 사용)"
    sheet["A1"].font = Font(bold=True, size=12)
    for column_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=3, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    for row_offset, row_values in enumerate(build_rows(mapping), start=4):
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_offset, column=column_index, value=value)
    widths = [6, 28, 30, 55, 45, 20]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(ord("A") + column_index - 1)].width = width

    # AI 초안 마커 (숨김 메타 시트 — §4.3-6 세탁 경로 차단)
    meta = workbook.create_sheet(DRAFT_META_SHEET)
    meta.sheet_state = "hidden"
    for row_index, (key, value) in enumerate([
        ("generated_by", f"{SKILL_NAME} v{SKILL_VERSION} (모드 B)"),
        ("generated_at", datetime.now(tz=timezone.utc).astimezone().isoformat()),
        ("mapping_revision", str(mapping["gate"].get("revision"))),
        ("reviewed", "false"),
    ], start=1):
        meta.cell(row=row_index, column=1, value=key)
        meta.cell(row=row_index, column=2, value=value)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    logger.info("Target 시트 초안 생성 (%d행, 마커 포함) → %s",
                len(build_rows(mapping)), output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
