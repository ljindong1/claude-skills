"""체크리스트 Target 시트를 파싱해 점검 대상 산출물 항목 목록(JSON)을 만든다.

사용: python parse_target_sheet.py <checklist.xlsx> [--sheet NAME] -o items.json
구조가 예상과 다르면 추측하지 않고 오류로 중단한다 (references/target_sheet.md).
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path

import openpyxl

from config import (
    GUIDE_ROW_MARK,
    HEADER_FILE_HINT,
    HEADER_GROUP_HINT,
    HEADER_NO,
    HEADER_NOTE_HINT,
    HEADER_REMARK_HINT,
    HEADER_WORK_PRODUCT,
    NON_TARGET_MARK,
    SHEET_PREFIX,
    TARGET_MARK,
)

logger = logging.getLogger(__name__)


class TargetSheetError(Exception):
    """Target 시트 구조가 예상과 달라 파싱할 수 없는 경우."""


@dataclass(frozen=True, slots=True)
class ColumnMap:
    """헤더 탐지로 확정한 열 인덱스(0-base)."""

    no: int
    group: int
    work_product: int
    files: int
    remark: int
    note: int | None


@dataclass(slots=True)
class TargetItem:
    """Target 시트의 산출물 항목 1건."""

    no: int
    procedure_group: str
    work_product: str
    listed_entries: list[str] = field(default_factory=list)
    target: bool = True
    remark: str = ""
    note: str = ""


def find_target_sheet(workbook: openpyxl.Workbook, sheet_name: str | None) -> str:
    """Target 시트명을 찾는다. 지정명이 있으면 검증만 한다."""
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise TargetSheetError(f"지정한 시트가 없음: {sheet_name}")
        return sheet_name
    for name in workbook.sheetnames:
        if name.startswith(SHEET_PREFIX):
            return name
    raise TargetSheetError(
        f"'{SHEET_PREFIX}'로 시작하는 시트를 찾지 못함 — 시트 목록: {workbook.sheetnames}"
    )


def detect_columns(rows: list[tuple[object, ...]]) -> tuple[int, ColumnMap]:
    """헤더 행을 탐지해 (헤더 행 인덱스, 열 매핑)을 돌려준다."""
    for row_index, row in enumerate(rows[:10]):
        texts = [str(cell).strip() if cell is not None else "" for cell in row]
        if HEADER_NO in texts and any(HEADER_WORK_PRODUCT in t for t in texts):
            no_col = texts.index(HEADER_NO)
            wp_col = next(i for i, t in enumerate(texts) if HEADER_WORK_PRODUCT in t)
            group_col = next(
                (i for i, t in enumerate(texts) if HEADER_GROUP_HINT in t and i != wp_col),
                no_col + 1,
            )
            files_col = next(
                (i for i, t in enumerate(texts) if HEADER_FILE_HINT in t), wp_col + 1
            )
            remark_col = next(
                (i for i, t in enumerate(texts) if HEADER_REMARK_HINT in t), files_col + 1
            )
            note_col = next(
                (i for i, t in enumerate(texts) if t.lower() == HEADER_NOTE_HINT), None
            )
            return row_index, ColumnMap(
                no=no_col, group=group_col, work_product=wp_col,
                files=files_col, remark=remark_col, note=note_col,
            )
    raise TargetSheetError(
        f"헤더 행 탐지 실패 — '{HEADER_NO}'와 '{HEADER_WORK_PRODUCT}'가 함께 있는 행 없음"
    )


def split_file_entries(cell_value: str) -> list[str]:
    """파일명 셀을 개행으로 분리하고 불릿(-)·공백을 정리한다."""
    entries: list[str] = []
    for line in cell_value.splitlines():
        token = line.strip().lstrip("-").strip().rstrip("/").strip()
        if token:
            entries.append(token)
    return entries


def parse_items(rows: list[tuple[object, ...]], columns: ColumnMap, header_row: int
                ) -> tuple[list[TargetItem], list[dict[str, object]]]:
    """헤더 이후 데이터 행을 항목으로 변환한다.

    v0.3: 절차그룹 병합 셀 forward-fill, 해석 불가 행은 UNPARSEABLE_ROWS 로 분리
    (items 미포함 — 커버리지 분모 제외, 대시보드 경고 노출).
    """
    items: list[TargetItem] = []
    unparseable: list[dict[str, object]] = []
    current_group = ""
    for row_number, row in enumerate(rows[header_row + 1:], start=header_row + 2):
        cells = [str(cell).strip() if cell is not None else "" for cell in row]

        def cell_at(index: int | None) -> str:
            return cells[index] if index is not None and index < len(cells) else ""

        first = cell_at(0)
        if first.startswith(GUIDE_ROW_MARK) or first.startswith("<작성"):
            break
        work_product = cell_at(columns.work_product)
        if not work_product:
            continue
        group = cell_at(columns.group) or current_group  # forward-fill
        current_group = group
        no_text = cell_at(columns.no)
        try:
            item_no = int(float(no_text))
        except ValueError:
            unparseable.append({
                "row": row_number,
                "reason": f"No. 열이 숫자가 아님: {no_text!r}",
                "work_product": work_product,
            })
            continue
        remark = cell_at(columns.remark)
        is_target = TARGET_MARK in remark and NON_TARGET_MARK not in remark
        items.append(
            TargetItem(
                no=item_no,
                procedure_group=group,
                work_product=work_product,
                listed_entries=split_file_entries(cell_at(columns.files)),
                target=is_target,
                remark=remark,
                note=cell_at(columns.note) if columns.note is not None else "",
            )
        )
    if not items:
        raise TargetSheetError("데이터 행 0건 — 시트 구조 확인 필요")
    # v0.4 DUPLICATE_ITEM_NO (FATAL): T-id 유일성은 계약의 근간 — 자동 재부여 금지
    seen: dict[int, int] = {}
    duplicates: list[str] = []
    for item in items:
        if item.no in seen:
            duplicates.append(f"No.{item.no} ('{items[seen[item.no]].work_product}' ↔ '{item.work_product}')")
        else:
            seen[item.no] = items.index(item)
    if duplicates:
        raise TargetSheetError(
            "DUPLICATE_ITEM_NO — Target 시트 No. 중복으로 안정 식별자(T-id)를 만들 수 없음. "
            f"시트 정비 후 재실행 필요 (자동 재부여 금지 — 계약 규칙 2): {'; '.join(duplicates)}"
        )
    return items, unparseable


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Target 시트 파싱")
    parser.add_argument("checklist", help="체크리스트 .xlsx 경로")
    parser.add_argument("--sheet", default=None, help="Target 시트명 (미지정 시 자동 탐지)")
    parser.add_argument("-o", "--output", required=True, help="items.json 출력 경로")
    args = parser.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")

    checklist_path = Path(args.checklist)
    if not checklist_path.is_file():
        logger.error("체크리스트 파일 없음: %s", checklist_path)
        return 1
    try:
        workbook = openpyxl.load_workbook(checklist_path, read_only=True, data_only=True)
        sheet_name = find_target_sheet(workbook, args.sheet)
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        header_row, columns = detect_columns(rows)
        items, unparseable = parse_items(rows, columns, header_row)
    except TargetSheetError as error:
        logger.error("파싱 중단: %s", error)
        return 1

    if unparseable:
        logger.warning("UNPARSEABLE_ROWS: %d행 해석 불가 — 경고로 기록", len(unparseable))
    result = {
        "source_file": checklist_path.name,
        "sheet": sheet_name,
        "items": [asdict(item) for item in items],
        "unparseable_rows": unparseable,
    }
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    target_count = sum(1 for item in items if item.target)
    logger.info(
        "파싱 완료: 항목 %d건 (점검 대상 %d, 미대상 %d) → %s",
        len(items), target_count, len(items) - target_count, output_path,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
