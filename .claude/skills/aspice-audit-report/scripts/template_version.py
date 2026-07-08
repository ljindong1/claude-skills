#!/usr/bin/env python3
"""체크리스트의 '양식 버전'(Template History 개정번호) 추출·비교.

양식 버전의 기준은 "Template History"(양식 개정 이력) 시트의 최대 개정번호다.
Cover의 문서이력(V1.0, V5.1 등)은 프로젝트 결과물의 문서 버전이므로 쓰지 않는다.

usage:
  python template_version.py <체크리스트.xlsx>                        # 개정번호·레이아웃 지문 출력
  python template_version.py <체크리스트.xlsx> --against <template_version.json>  # 보유 버전과 비교

비교 판정:
  NEWER   — 첨부가 보유보다 최신 → 첨부 양식으로 작업 + 스킬 기준 양식 갱신(.skill 재패키징) 제안
  SAME    — 동일(문항 내용까지 일치) → 보유 최신 양식으로 진행
  SAME·CONTENT_DIFF — 개정번호는 같은데 실문항이 다름(프로젝트 테일러링 가능성,
            예: JX HEV의 Process 122~127 추가·23문항 문구 상이) → 프로젝트 전용
            체크리스트인지 사용자에게 확인. 그 프로젝트 점검이면 '첨부'가 기준이다
  OLDER   — 첨부가 구버전 → 보유 최신 양식으로 진행(첨부는 증분 근거로만)
  UNKNOWN — Template History 없음 → 레이아웃 지문 대조. LAYOUT_SAME이면 동일 양식으로
            간주 가능, LAYOUT_DIFF면 어느 쪽이 최신인지 사용자에게 확인(임의 단정 금지)
종료코드: 0=SAME/OLDER, 2=NEWER, 3=UNKNOWN, 4=SAME·CONTENT_DIFF (자동 분기용)
"""
import sys
import json
import hashlib
import argparse
import openpyxl
import config

HISTORY_SHEET = "Template History"


def read_revision(wb):
    """Template History 시트에서 (최대 개정번호, 그 행의 제·개정일자) 반환. 없으면 (None, None)."""
    if HISTORY_SHEET not in wb.sheetnames:
        return None, None
    ws = wb[HISTORY_SHEET]
    best, date = None, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[0] if row else None
        if v is None:
            continue
        t = str(v).strip()
        if t.isdigit():
            r = int(t)
            if best is None or r > best:
                best, date = r, (str(row[1]).strip() if len(row) > 1 and row[1] else None)
    return best, date


def fingerprint(wb):
    """레이아웃 지문(머리글) + 문항 내용 지문(문항 수·텍스트 해시).

    - layout: 시트 종류별 3행 머리글 목록 → 개정번호 없는 파일의 보조 판별용
    - content: 시트 종류별 문항 수와 문항 텍스트 sha1 → 같은 개정번호라도
      프로젝트 테일러링(문항 추가/문구 수정)이 있는지 판별용
    """
    sh = config.find_sheets(wb)
    layout, content = {}, {}
    for kind in ("process", "wp", "target"):
        if sh[kind]:
            ws = wb[sh[kind][0]]
            hdr = []
            for c in range(1, (ws.max_column or 0) + 1):
                v = ws.cell(config.HEADER_ROW, c).value
                if v is not None and str(v).strip():
                    hdr.append(str(v).strip())
            layout[kind] = hdr
            if kind in ("process", "wp"):
                C = config.detect_cols(ws, kind)
                qs = []
                for r in range(config.DATA_START_ROW, ws.max_row + 1):
                    q = ws.cell(r, C["question"]).value
                    if q and str(q).strip():
                        qs.append(" ".join(str(q).split()))
                content[kind] = {"n": len(qs),
                                 "hash": hashlib.sha1("\n".join(qs).encode()).hexdigest()[:12]}
        else:
            layout[kind] = None
    digest = hashlib.sha1(json.dumps(layout, ensure_ascii=False, sort_keys=True).encode()).hexdigest()[:12]
    return digest, layout, content


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--against", help="보유 버전 기록(template_version.json)")
    a = ap.parse_args()
    wb = openpyxl.load_workbook(a.file, read_only=False, data_only=True)
    rev, date = read_revision(wb)
    fp, layout, content = fingerprint(wb)
    print(f"첨부 양식: 개정번호={rev if rev is not None else '불명(Template History 없음)'}"
          f"{' (' + date + ')' if date else ''} | 레이아웃 지문={fp}")
    print(f"  시트 구성: " + ", ".join(k for k, v in layout.items() if v))
    if content:
        print("  문항: " + ", ".join(f"{k} {v['n']}개({v['hash']})" for k, v in content.items()))
    if not a.against:
        print(json.dumps({"revision": rev, "revised": date, "fingerprint": fp,
                          "content": content}, ensure_ascii=False))
        return 0
    with open(a.against, encoding="utf-8") as f:
        ref = json.load(f)
    print(f"보유 양식: 개정번호={ref.get('revision')} ({ref.get('revised')}) | 지문={ref.get('fingerprint')}")
    if rev is not None and ref.get("revision") is not None:
        if rev > ref["revision"]:
            print("[NEWER] 첨부가 더 최신 양식 → 첨부 양식으로 작업하고, 스킬 기준 양식 갱신(.skill 재패키징)을 제안하라")
            return 2
        if rev == ref["revision"]:
            refc = ref.get("content") or {}
            diff = [k for k in content
                    if k in refc and (content[k]["n"] != refc[k]["n"] or content[k]["hash"] != refc[k]["hash"])]
            if diff:
                print(f"[SAME·CONTENT_DIFF] 개정번호는 같지만 실문항이 다름({', '.join(diff)}) — "
                      "프로젝트 테일러링 가능성. 이 프로젝트의 점검이면 '첨부 체크리스트'가 기준이다. "
                      "양식 표준 변경인지 프로젝트 전용인지 사용자/담당자에게 확인하라")
                return 4
            print("[SAME] 동일 양식 버전(문항 내용 일치) → 보유 최신 양식으로 진행")
            return 0
        print("[OLDER] 첨부가 구버전 → 보유 최신 양식으로 진행(첨부는 증분/직전 차수 근거로만 사용)")
        return 0
    if fp == ref.get("fingerprint"):
        print("[UNKNOWN·LAYOUT_SAME] 개정번호 없음, 레이아웃 동일 → 동일 양식으로 간주 가능")
        return 3
    print("[UNKNOWN·LAYOUT_DIFF] 개정번호 없음 + 레이아웃 상이 → 어느 쪽이 최신인지 사용자에게 확인(임의 단정 금지)")
    return 3


if __name__ == "__main__":
    sys.exit(main())
