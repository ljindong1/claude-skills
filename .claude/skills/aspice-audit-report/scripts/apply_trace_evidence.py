#!/usr/bin/env python3
"""추적성 체커 결과(trace_evidence.json)를 채워진 결과 체크리스트에 반영한다.

추적/일관/할당 관련 WP 항목에 export 기반 미연결 건수를 근거로 붙인다.
- 활성 갭>0 → Result 비움(확인 필요) + "활성 미연결 N건, Fail 후보" 코멘트. (자동 Pass 금지)
- 활성 갭=0 → 해당 항목 Result 비어있으면 Pass + "추적성 완전(활성 미연결 0)".

usage: python apply_trace_evidence.py <결과.xlsx> <trace_evidence.json>
"""
import sys
import re
import json
import openpyxl
import config

TRACE_KW = ['추적', '일관', '할당']
# WP 산출물(정규화 부분문자열) → 관련 trace 파일 키
DELIV_PAIRS = [
    ('시스템요구사항', ['SR_SysRS', 'SysRS_SysAD']),
    ('systemrequirements', ['SR_SysRS', 'SysRS_SysAD']),
    ('소프트웨어요구사양서', ['SWR_SWA', 'SWR-SysAD']),
    ('소프트웨어요구사항', ['SWR_SWA', 'SWR-SysAD']),
    ('소프트웨어아키텍처', ['SWR_SWA', 'SWA_SWD']),
    ('소프트웨어아키텍쳐', ['SWR_SWA', 'SWA_SWD']),
    ('소프트웨어상세설계', ['SWA_SWD']),
]


def norm(s):
    return re.sub(r'[\s()◆\[\]]', '', str(s or '')).lower()


def main(result, trace_json):
    with open(trace_json, encoding='utf-8') as f:
        ev = json.load(f)
    bykey = {}
    for e in ev:
        for k in ('SR_SysRS', 'SysRS_SysAD', 'SWR_SWA', 'SWA_SWD', 'SWR-SysAD'):
            if k.replace('-','').replace('_','').lower() in e['file'].replace('-','').replace('_','').lower():
                bykey[k] = e
    wb = openpyxl.load_workbook(result)
    sh = config.find_sheets(wb)
    W = config.WP_COLS
    ws = wb[sh['wp']]
    updated = 0
    for row in range(config.DATA_START_ROW, ws.max_row + 1):
        no = ws.cell(row, 1).value
        q = ws.cell(row, W['question']).value
        if no is None or not q:
            continue
        if not any(k in str(q) for k in TRACE_KW):
            continue
        prod = norm(ws.cell(row, W['product']).value)
        pairs = next((ks for sub, ks in DELIV_PAIRS if sub in prod), None)
        if not pairs:
            continue
        rel = [bykey[k] for k in pairs if k in bykey]
        if not rel:
            continue
        gaps = sum(e['gaps_active'] for e in rel)
        labels = ', '.join(f"{e['pair']}:{e['gaps_active']}" for e in rel)
        if gaps > 0:
            ws.cell(row, W['applicable']).value = 'Yes'
            ws.cell(row, W['result']).value = None  # 자동 Pass 금지
            ws.cell(row, W['comments']).value = (
                f"260623 (AI 1차·추적성): export 기준 활성 미연결 {gaps}건 ({labels}). "
                f"Fail 후보 — 미할당/미추적 사유 확인 필요")
        else:
            ws.cell(row, W['applicable']).value = 'Yes'
            if not ws.cell(row, W['result']).value:
                ws.cell(row, W['result']).value = 'Pass'
            ws.cell(row, W['comments']).value = (
                f"260623 (AI 1차·추적성): export 기준 활성 미연결 0건 ({labels}). 추적성 완전")
        updated += 1
    wb.save(result)
    print(f"추적성 반영 행: {updated}  (사용 pair: {sorted(bykey)})")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(__doc__); sys.exit(1)
    main(sys.argv[1], sys.argv[2])
