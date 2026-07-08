#!/usr/bin/env python3
"""사람 완성본 → 빈 양식. 답 칸만 비우고 문항·구조·드롭다운은 보존한다.

usage:
  # 전체 클리어(완전 새 점검): 인식된 모든 시트의 답 칸을 비운다
  python make_blank_template.py <완성본.xlsx> <빈양식.xlsx>

  # 증분 — 지정 시트만 클리어(재점검): 다른 시트의 기존 결과는 그대로 보존
  python make_blank_template.py <완성본.xlsx> <출력.xlsx> --sheet "Process Checklist_소프트웨어 설계 단계말"

  # 증분 — 시트 복사 추가(새 차수, 권장): 기존 시트를 복사해 새 차수 시트를 만들고
  # 그 시트만 비운다. 기존 완료 시트는 전부 보존(직전 차수 스코핑 근거로 활용).
  python make_blank_template.py <완성본.xlsx> <출력.xlsx> --new-sheet "Process Checklist_테스트 단계말" --copy-from "Process Checklist_소프트웨어 설계 단계말"
"""
import sys
import argparse
import openpyxl
from openpyxl.cell.cell import MergedCell
import config


def clear_sheet(ws, kind):
    cols = config.detect_cols(ws, kind)
    if kind in ("process", "wp") and not cols.get("applicable"):
        # 구버전 레이아웃 — 신규 작성 금지(읽기 전용 보존). 최신 양식으로 업데이트 필요.
        return None, cols
    cleared = 0
    for row in range(config.DATA_START_ROW, ws.max_row + 1):
        for c in config.clear_cols(cols, kind):
            cell = ws.cell(row=row, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                cell.value = None
                cleared += 1
            if cell.hyperlink is not None:
                # 하이퍼링크가 남으면 저장 시 URL이 값으로 되살아난다(부적합 이슈 링크 실측)
                cell.hyperlink = None
    return cleared, cols


def main(src, out, only_sheets=None, new_sheet=None, copy_from=None, keep_images=False):
    wb = openpyxl.load_workbook(src)  # 서식/드롭다운 보존 (data_only 금지)
    sheets = config.find_sheets(wb)
    all_names = [n for names in sheets.values() for n in names]
    work = []  # (시트명, kind) — 클리어 대상

    if new_sheet:
        if not copy_from:
            sys.exit("--new-sheet 에는 --copy-from <원본시트명> 이 필요합니다")
        if copy_from not in wb.sheetnames:
            sys.exit(f"원본 시트 없음: {copy_from!r} (있는 시트: {wb.sheetnames})")
        if new_sheet in wb.sheetnames:
            sys.exit(f"이미 존재하는 시트명: {new_sheet!r}")
        kind = config.sheet_kind(new_sheet) or config.sheet_kind(copy_from)
        if not kind:
            sys.exit(f"시트 종류 인식 불가(접두어 확인): {new_sheet!r}")
        ns = wb.copy_worksheet(wb[copy_from])
        ns.title = new_sheet
        work = [(new_sheet, kind)]
        print(f"시트 복사: {copy_from!r} → {new_sheet!r} (기존 시트 전부 보존)")
    elif only_sheets:
        for name in only_sheets:
            if name not in wb.sheetnames:
                sys.exit(f"시트 없음: {name!r} (있는 시트: {wb.sheetnames})")
            kind = config.sheet_kind(name)
            if not kind:
                sys.exit(f"체크리스트 시트가 아님(접두어 확인): {name!r}")
            work.append((name, kind))
        keep = [n for n in all_names if n not in only_sheets]
        if keep:
            print(f"보존되는 기존 시트: {keep}")
    else:
        work = [(n, k) for k, names in sheets.items() for n in names]
        # 전체 클리어 = 새 양식/새 프로젝트 용도 → 완성본의 증적 캡처 이미지는 제거
        # (실측: JX HEV 완성본에 증적 PNG 97장, 15.5MB → 제거 후 0.1MB. 문항·서식 영향 없음)
        if not keep_images:
            n_img = 0
            for ws in wb.worksheets:
                n_img += len(ws._images)
                ws._images = []
            if n_img:
                print(f"  증적 이미지 제거: {n_img}장 (--keep-images로 보존 가능)")

    if not work:
        sys.exit("클리어할 체크리스트 시트를 찾지 못했습니다: " + str(wb.sheetnames))

    total = 0
    skipped = []
    for name, kind in work:
        n, cols = clear_sheet(wb[name], kind)
        if n is None:
            skipped.append(name)
            print(f"  [스킵] {name!r}: 구버전 레이아웃(Applicable 열 없음) — 신규 작성 금지, 읽기 전용 보존. 최신 양식(assets/latest_template.xlsx)으로 작업하라")
            continue
        total += n
        print(f"  클리어: {name!r} ({kind}) 답 칸 {n}개")
    if skipped and total == 0:
        sys.exit("클리어 가능한 최신 레이아웃 시트가 없습니다 — 구버전 양식. 스킬 보유 최신 양식으로 진행하라")
    wb.save(out)
    print(f"빈 양식 저장: {out} (비운 답 칸 {total}개)")

    # 컬럼 검증용 한 행 출력 — 문항이 보존돼야 정상
    for name, kind in work:
        if kind == "process":
            ws = wb[name]
            cols = config.detect_cols(ws, kind)
            q = ws.cell(config.DATA_START_ROW, cols["question"]).value
            print(f"  [검증] {name!r} {config.DATA_START_ROW}행 문항 칸 = {str(q)[:50]!r}")
            break


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("src")
    ap.add_argument("out")
    ap.add_argument("--sheet", action="append", help="이 시트만 클리어(반복 지정 가능)")
    ap.add_argument("--new-sheet", help="복사로 추가할 새 시트명")
    ap.add_argument("--copy-from", help="--new-sheet의 원본 시트명")
    ap.add_argument("--keep-images", action="store_true", help="전체 클리어 시에도 이미지 보존")
    a = ap.parse_args()
    main(a.src, a.out, a.sheet, a.new_sheet, a.copy_from, a.keep_images)
