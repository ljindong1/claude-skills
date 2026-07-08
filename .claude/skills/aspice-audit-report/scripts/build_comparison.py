#!/usr/bin/env python3
"""사람 작성본 vs AI 리포트 코멘트 비교표(검증 산출물) — 시트별 분리.

최종 확인 단계에서 사람 코멘트(좌)와 AI 코멘트(우)를 항목별로 나란히 두어
사람급 도달 여부를 사람이 눈으로 대조하게 한다.

★ 검증용 임시 산출물이다. 파일명에 _검증용을 붙여 최종 납품물(체크리스트)과 구분하고
   검증이 끝나면 삭제해도 된다.

- 양식에 있는 WP/Process 시트를 전부(같은 이름 시트끼리) 비교한다.
  WP 없는 Process 전용 양식이면 Process만.
- --sheet 지정 시 그 시트만(증분 모드).
- Applicable 열 없는 구버전 시트는 Result 존재로 Yes를 추정.

usage:
  python build_comparison.py <결과.xlsx> <사람작성본.xlsx> <출력.xlsx> [--sheet 시트명 ...]
"""
import argparse
import openpyxl
import config
from collections import defaultdict
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DEEP = ["AI·본문", "AI·결과서", "AI·검토", "AI·정독", "AI·구조", "AI·추적",
        "AI·데이터", "AI·확인", "본문+검토", "본문정독", "검토결과서"]


def is_deep(c):
    return bool(c) and any(m in str(c) for m in DEEP)


def rows(ws, kind):
    C = config.detect_cols(ws, kind)
    gcol = C["product"] if kind == "wp" else C["output_product"]
    apc, rc = C.get("applicable"), C["result"]
    cur = None
    for r in range(config.DATA_START_ROW, ws.max_row + 1):
        no = ws.cell(r, 1).value
        g = ws.cell(r, gcol).value
        if g and str(g).strip():
            cur = str(g).strip().splitlines()[0][:28]
        if no is None:
            continue
        if apc:
            ap = str(ws.cell(r, apc).value or "").strip()
        else:
            ap = "Yes" if str(ws.cell(r, rc).value or "").strip() else "No"
        yield (no, cur, ws.cell(r, C["question"]).value, ap, ws.cell(r, C["comments"]).value)


def style_header(rowcells, hdr, fill, bd):
    for c in rowcells:
        c.font = hdr; c.fill = fill
        c.alignment = Alignment(horizontal="center"); c.border = bd


def main(result, human, out, only=None):
    awb = openpyxl.load_workbook(result, data_only=True)
    hwb = openpyxl.load_workbook(human, data_only=True)
    sh = config.find_sheets(awb)

    hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="305496")
    thin = Side(style="thin", color="BBBBBB"); bd = Border(thin, thin, thin, thin)

    owb = openpyxl.Workbook()
    summ = owb.active; summ.title = "요약"
    summ.append(["시트", "산출물/활동", "사람 코멘트(Yes)", "AI 코멘트", "AI 사람급(심층)", "비고"])
    style_header(summ[1], hdr, fill, bd)

    made = 0
    for kind in ("wp", "process"):
        for name in sh[kind]:
            if only and name not in only:
                continue
            if name not in hwb.sheetnames:
                continue
            aws, hws = awb[name], hwb[name]
            ai_pk = {}; ai_no = {}
            for no, g, q, ap, c in rows(aws, kind):
                ai_pk[(g, no)] = c; ai_no[no] = c

            ps = defaultdict(lambda: [0, 0, 0]); order = []; detail = []
            for no, g, q, ap, hc in rows(hws, kind):
                if ap != "Yes" or not (hc and str(hc).strip()):
                    continue
                if g not in ps:
                    order.append(g)
                ac = ai_pk.get((g, no)) if kind == "wp" else ai_no.get(no)
                ps[g][0] += 1
                if ac and str(ac).strip():
                    ps[g][1] += 1
                    if is_deep(ac):
                        ps[g][2] += 1
                detail.append((no, g, q, hc, ac))

            for g in order:
                a, b, c = ps[g]
                summ.append([name, g, a, b, c, "사람급 달성" if c == a else "보완 필요"])

            ds = owb.create_sheet(("비교_" + name)[:31])
            col2 = "산출물" if kind == "wp" else "활동(출력 산출물)"
            ds.append(["No", col2, "점검 항목", "사람 Review Comments", "AI Review Comments"])
            style_header(ds[1], hdr, fill, bd)
            detail.sort(key=lambda x: (order.index(x[1]) if x[1] in order else 999, x[0]))
            for no, g, q, hc, ac in detail:
                ds.append([no, g, str(q or "").strip(), str(hc).strip(), str(ac or "").strip()])
            for row in ds.iter_rows(min_row=2):
                for c in row:
                    c.border = bd; c.alignment = Alignment(vertical="top", wrap_text=True)
            ds.column_dimensions["A"].width = 6; ds.column_dimensions["B"].width = 24
            ds.column_dimensions["C"].width = 32; ds.column_dimensions["D"].width = 60
            ds.column_dimensions["E"].width = 60
            ds.freeze_panes = "A2"
            made += 1

    if not made:
        raise SystemExit("비교할 공통 시트가 없습니다")

    for row in summ.iter_rows(min_row=2):
        for c in row:
            c.border = bd; c.alignment = Alignment(vertical="center", wrap_text=True)
    summ.column_dimensions["A"].width = 30; summ.column_dimensions["B"].width = 26
    for col in "CDE":
        summ.column_dimensions[col].width = 13
    summ.column_dimensions["F"].width = 16
    summ.freeze_panes = "A2"

    owb.save(out)
    print(f"비교표 저장: {out}  (시트: {[s.title for s in owb.worksheets]})")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("result")
    ap.add_argument("human")
    ap.add_argument("out")
    ap.add_argument("--sheet", action="append", help="이 시트만 비교(증분 모드)")
    a = ap.parse_args()
    main(a.result, a.human, a.out, a.sheet)
