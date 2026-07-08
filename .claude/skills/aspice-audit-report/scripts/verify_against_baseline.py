#!/usr/bin/env python3
"""AI 작성본 vs 사람 작성본 행 단위 비교 → 검증표(요약 + 불일치 상세).

usage:
  python verify_against_baseline.py <AI작성본.xlsx> <사람작성본.xlsx> <검증표.xlsx>
  python verify_against_baseline.py <AI.xlsx> <사람.xlsx> <검증표.xlsx> --sheet "Process Checklist_테스트 단계말"

두 파일은 같은 양식이어야 한다(같은 이름 시트끼리 행 1:1 비교).
- 시트는 두 파일에 공통으로 존재하는 것을 전부 비교한다(WP 없으면 Process만).
- Applicable 열이 없는 구버전 시트는 Result 존재 여부로 Yes/No를 추정한다.
- Pass/Conditional pass는 '합격 계열'로 같은 묶음 취급(config.PASS_FAMILY).
"""
import re
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config

PASS_TAG_RE = re.compile(r"판정:\s*Pass(?!\s*후보)")


def tag_of(c):
    t = str(c or "")
    if PASS_TAG_RE.search(t):
        return "Pass"
    if "Pass 후보" in t:
        return "Pass후보"
    if "Fail 후보" in t:
        return "Fail후보"
    return 


def n(v):
    return "" if v is None else str(v).strip()


def collect(H, A, sheet, kind):
    h, a = H[sheet], A[sheet]
    C = config.detect_cols(h, kind)
    qcol = C["question"]
    labcol = C["product"] if kind == "wp" else C["output_product"]
    apc, rc = C.get("applicable"), C["result"]

    def ap_of(ws, r):
        if apc:
            return n(ws.cell(r, apc).value)
        return "Yes" if n(ws.cell(r, rc).value) else "No"  # 구버전: Result로 추정

    cc = C["comments"]
    out = []
    for r in range(config.DATA_START_ROW, h.max_row + 1):
        q = h.cell(r, qcol).value
        if not q:
            continue
        out.append((r, n(h.cell(r, labcol).value), n(q),
                    ap_of(h, r), n(h.cell(r, rc).value),
                    ap_of(a, r), n(a.cell(r, rc).value),
                    tag_of(a.cell(r, cc).value)))
    return out


def eff(v):
    """'확인 필요'는 판정이 아니라 보류 표기 — 비교에서 미판정(빈값)으로 취급."""
    return "" if str(v or "").strip() in ("확인 필요", "확인필요") else v


def pf(v):
    return "PASS계열" if v in config.PASS_FAMILY else v


def classify(ha, hr, aa, ar):
    ar = eff(ar)
    hy, ay = ha == "Yes", aa == "Yes"
    if hy and ay:
        if not ar:
            return "미판정"
        if hr == "Fail" and ar in config.PASS_FAMILY:
            return "위험오판"
        if pf(hr) == pf(ar):
            return "일치"
        return "결과불일치"
    if ay and not hy:
        return "과대"
    if hy and not ay:
        return "과소"
    return "일치"


def metrics(rows):
    m = dict(tot=len(rows), scope=0, under=0, over=0, rden=0, rmatch=0, rmiss=0, fp=0,
             prop_n=0, prop_ok=0, prop_fp=0)
    for r, lab, q, ha, hr, aa, ar, atag in rows:
        ar = eff(ar)
        # 제안(태그) 정확도 — Result를 비워두는 '후보' 정책과 무관하게 제안 품질을 추적
        if ha == "Yes" and hr and atag:
            m["prop_n"] += 1
            if (atag in ("Pass", "Pass후보") and hr in config.PASS_FAMILY) or (atag == "Fail후보" and hr == "Fail"):
                m["prop_ok"] += 1
            elif atag in ("Pass", "Pass후보") and hr == "Fail":
                m["prop_fp"] += 1
        hy, ay = ha == "Yes", aa == "Yes"
        if ha == aa:
            m["scope"] += 1
        elif hy and not ay:
            m["under"] += 1
        elif ay and not hy:
            m["over"] += 1
        if hy and hr:
            m["rden"] += 1
            if ay and not ar:
                m["rmiss"] += 1
            elif ay and hr == "Fail" and ar in config.PASS_FAMILY:
                m["fp"] += 1
            elif ay and pf(hr) == pf(ar):
                m["rmatch"] += 1
            elif not ay:
                m["rmiss"] += 1
    return m


def main(ai, human, out, only=None):
    A = openpyxl.load_workbook(ai, data_only=True)
    H = openpyxl.load_workbook(human, data_only=True)
    sh = config.find_sheets(H)

    sections = []  # (라벨=시트명, kind, rows)
    for kind in ("wp", "process"):
        for name in sh[kind]:
            if only and name not in only:
                continue
            if name not in A.sheetnames:
                print(f"[주의] AI 작성본에 시트 없음 → 건너뜀: {name!r}")
                continue
            sections.append((name, kind, collect(H, A, name, kind)))
    if not sections:
        raise SystemExit("비교할 공통 시트가 없습니다")

    det = []
    for label, kind, rows in sections:
        for r, lab, q, ha, hr, aa, ar, atag in rows:
            cat = classify(ha, hr, aa, ar)
            if cat in ("과소", "과대", "위험오판", "결과불일치"):
                det.append((cat, label, r, lab, q[:60],
                            f"{ha or '-'}/{hr or '-'}", f"{aa or '-'}/{ar or '-'}"))
    order = {"위험오판": 0, "과소": 1, "결과불일치": 2, "과대": 3}
    det.sort(key=lambda x: (order[x[0]], x[1], x[2]))
    mets = [(label, metrics(rows)) for label, kind, rows in sections]

    wb = openpyxl.Workbook()
    F, navy, band, warn = "맑은 고딕", "1F3864", "D9E1F2", "FCE4D6"
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(thin, thin, thin, thin)

    def st(c, b=False, sz=10, col="000000", fill=None, h="left"):
        c.font = Font(name=F, bold=b, size=sz, color=col)
        c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
        c.border = bd
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)

    s = wb.active
    s.title = "요약"
    s.merge_cells("A1:E1"); s["A1"] = "AI 1차 자동 점검 vs 사람 작성본 — 검증 요약"
    st(s["A1"], True, 14, "FFFFFF", navy, "center")
    s["A3"] = "[ ① 스코핑(Applicable Yes/No) 정확도 ]"; st(s["A3"], True, 11, "1F3864", band); s.merge_cells("A3:E3")
    for j, x in enumerate(["시트", "전체", "일치", "AI 과소(누락)", "AI 과대"]):
        st(s.cell(4, j + 1), True, 10, "FFFFFF", navy, "center"); s.cell(4, j + 1).value = x
    for i, (nm, m) in enumerate(mets):
        pct = round(m["scope"] / m["tot"] * 100) if m["tot"] else 0
        for j, v in enumerate([nm, m["tot"], f"{m['scope']} ({pct}%)", m["under"], m["over"]]):
            st(s.cell(5 + i, j + 1), False, 10, "000000", None, "center"); s.cell(5 + i, j + 1).value = v
    r2 = 6 + len(mets)
    s.cell(r2, 1).value = "[ ② 결과 판정 정확도 (사람이 판정한 항목 기준, Pass/Conditional pass=합격 계열) ]"
    st(s.cell(r2, 1), True, 11, "1F3864", band); s.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=5)
    for j, x in enumerate(["시트", "사람 판정 수", "AI 일치", "AI 미판정", "⚠⚠ 위험오판"]):
        st(s.cell(r2 + 1, j + 1), True, 10, "FFFFFF", navy, "center"); s.cell(r2 + 1, j + 1).value = x
    for i, (nm, m) in enumerate(mets):
        for j, v in enumerate([nm, m["rden"], m["rmatch"], m["rmiss"], m["fp"]]):
            st(s.cell(r2 + 2 + i, j + 1), False, 10, "C00000" if (j == 4 and v) else "000000", None, "center")
            s.cell(r2 + 2 + i, j + 1).value = v
    for col, w in zip("ABCDE", [34, 12, 16, 16, 12]):
        s.column_dimensions[col].width = w

    d = wb.create_sheet("불일치 상세")
    d.merge_cells("A1:G1"); d["A1"] = f"불일치 상세 {len(det)}건 (위험오판→AI누락→결과상이→AI과대)"
    st(d["A1"], True, 12, "FFFFFF", navy, "center")
    for j, x in enumerate(["구분", "시트", "행", "산출물/절차", "문항", "사람 A/R", "AI A/R"]):
        st(d.cell(3, j + 1), True, 10, "FFFFFF", navy, "center"); d.cell(3, j + 1).value = x
    cmap = {"위험오판": warn, "과소": "FFF2CC", "결과불일치": "FFF2CC", "과대": "F2F2F2"}
    disp = {"위험오판": "⚠⚠ 위험오판", "과소": "⚠ AI 누락", "결과불일치": "결과 상이", "과대": "AI 과대"}
    for i, (cat, sheet, row, lab, q, hv, av) in enumerate(det):
        for j, v in enumerate([disp[cat], sheet, row, lab, q, hv, av]):
            st(d.cell(4 + i, j + 1), j == 0, 9, "000000", cmap[cat] if j == 0 else None,
               "center" if j in (1, 2, 5, 6) else "left")
            d.cell(4 + i, j + 1).value = v
    for col, w in zip("ABCDEFG", [12, 24, 6, 26, 46, 16, 16]):
        d.column_dimensions[col].width = w
    d.freeze_panes = "A4"
    wb.save(out)
    print("검증표 저장:", out)
    for nm, m in mets:
        print(nm, m)
        if m["prop_n"]:
            print(f"  └ 제안(태그) 정확도: {m['prop_ok']}/{m['prop_n']} ({m['prop_ok']/m['prop_n']*100:.0f}%)"
                  + (f" | ⚠ Pass계열 제안인데 사람 Fail {m['prop_fp']}건" if m['prop_fp'] else ""))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("ai")
    ap.add_argument("human")
    ap.add_argument("out")
    ap.add_argument("--sheet", action="append", help="이 시트만 비교(증분 모드)")
    a = ap.parse_args()
    main(a.ai, a.human, a.out, a.sheet)
