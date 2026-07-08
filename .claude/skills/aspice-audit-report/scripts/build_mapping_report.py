#!/usr/bin/env python3
"""매핑 통합 리포트(한 파일) — 매핑 테이블 + 커버리지(파일/체크시트)를 한 장에.

기존 build_mapping_table.py(매핑 테이블+파일 커버리지)와 mapping_report.py(요약+파일
관점+체크시트 관점)는 '파일 사용여부' 시트가 중복이었다. 이 스크립트는 둘을 합쳐
중복 없이 한 파일로 출력한다.

시트:
  - 요약        : 파일 관점/체크시트 관점 통계
  - 매핑 테이블  : 산출물(키) ↔ 파일/URL ↔ 내용 근거 (deliverable_map.json)
  - 파일 관점    : 폴더 파일이 사용/미사용/잣대/추적성 (결과.xlsx Target 기준 — 단일 출처)
  - 체크시트 관점: 산출물(Target)에 파일이 연결됐나/시스템URL/없음/미대상

usage:
  python build_mapping_report.py <deliverable_map.json> <결과.xlsx> <대상산출물폴더> <매핑리포트.xlsx>
결과.xlsx 는 fill_checklist.py로 Target이 채워진 파일이어야 한다.
"""
import sys
import os
import re
import json
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config

REFERENCE = ['bj1bdc', 'jx1', 'mpe1', 'mpe2', '표준', '지침서', 'pamv']
F, NAVY, BAND, WARN, OK, GREY = '맑은 고딕', '1F3864', 'D9E1F2', 'FCE4D6', 'E2EFDA', 'F2F2F2'
THIN = Side(style='thin', color='BFBFBF'); BD = Border(THIN, THIN, THIN, THIN)


def tok(s):
    return re.sub(r'[\s_\-(),\[\]/·.]', '', os.path.splitext(str(s or ''))[0]).lower()


def shared(a, b, n=10):
    if len(a) < n:
        return bool(a) and a in b
    return any(a[i:i + n] in b for i in range(len(a) - n + 1))


def st(c, b=False, sz=10, col='000000', fill=None, h='left'):
    c.font = Font(name=F, bold=b, size=sz, color=col)
    c.alignment = Alignment(horizontal=h, vertical='center', wrap_text=True); c.border = BD
    if fill:
        c.fill = PatternFill('solid', fgColor=fill)


def main(mappath, result, folder, out):
    ev = json.load(open(mappath, encoding='utf-8')).get('evidence', {})
    wb = openpyxl.load_workbook(result, data_only=True)
    # find_sheets는 종류별 '리스트'를 반환한다. Target 시트가 여러 개면 전부,
    # 없으면(Process 전용 양식) 매핑(evidence) 기준으로만 리포트를 만든다.
    targets = []
    for tname in config.find_sheets(wb)['target']:
        tws = wb[tname]
        T = config.detect_cols(tws, 'target')
        for r in range(config.DATA_START_ROW, tws.max_row + 1):
            prod = tws.cell(r, T['product']).value
            if not prod:
                continue
            f = str(tws.cell(r, T['file']).value or '')
            b = str(tws.cell(r, T['bigo']).value or '')
            targets.append((str(prod), f, b))
    if not targets:
        # Target 시트 부재 — deliverable_map의 evidence를 대체 소스로 사용
        print('[안내] Target 시트 없음 — 매핑(deliverable_map) 기준으로 리포트 생성')
        for kw, val in ev.items():
            fn = val[0] if isinstance(val, (list, tuple)) and val else str(val)
            targets.append((str(kw), str(fn), '[매핑 기준]'))

    # 체크시트 관점
    chk = []
    for prod, f, b in targets:
        if '미대상' in b:
            stt = '미대상'
        elif 'http' in f or ('시스템' in b and f):
            stt = '시스템(URL)'
        elif f and 'http' not in f and '미확인' not in f and '확인 필요' not in b:
            stt = '파일 연결됨'
        else:
            stt = '파일 없음 / 확인필요'
        chk.append((prod, stt, f or b))

    # 파일 관점 (결과 Target에 연결된 파일명 토큰 기준)
    allcells = [tok(c) for _, c, _ in targets if c]
    entries = os.listdir(folder)
    files = sorted([x for x in entries if os.path.isfile(os.path.join(folder, x))]
                   + [x + '/' for x in entries
                      if os.path.isdir(os.path.join(folder, x)) and not x.startswith('_')])
    fileside = []
    for fn in files:
        ft = tok(fn)
        if 'traceability' in ft:
            cat = '사용(추적성)'
        elif any(rk in ft for rk in REFERENCE):
            cat = '잣대/참조(점검대상 아님)'
        elif any(shared(ft, c) for c in allcells):
            cat = '사용됨'
        else:
            cat = '미사용'
        fileside.append((fn, cat))

    wbo = openpyxl.Workbook()
    # ---- 요약 ----
    s = wbo.active; s.title = '요약'
    s.merge_cells('A1:C1'); s['A1'] = '매핑 통합 리포트 (매핑 + 커버리지)'
    st(s['A1'], True, 14, 'FFFFFF', NAVY, 'center')
    fc = Counter(c for _, c in fileside); cc = Counter(x[1] for x in chk)
    s.merge_cells('A3:C3'); s['A3'] = '[ 파일 관점 — 폴더 파일이 쓰였나 ]'; st(s['A3'], True, 11, '1F3864', BAND)
    r = 4
    for k in ['사용됨', '사용(추적성)', '잣대/참조(점검대상 아님)', '미사용']:
        st(s.cell(r, 1)); s.cell(r, 1).value = k
        st(s.cell(r, 2), False, 10, 'C00000' if k == '미사용' and fc.get(k) else '000000', None, 'center')
        s.cell(r, 2).value = fc.get(k, 0); r += 1
    st(s.cell(r, 1), True, 10, fill=OK); s.cell(r, 1).value = f'합계 {len(fileside)}'; r += 2
    s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    s.cell(r, 1).value = '[ 체크시트 관점 — 산출물에 파일이 있나 ]'; st(s.cell(r, 1), True, 11, '1F3864', BAND); r += 1
    for k in ['파일 연결됨', '시스템(URL)', '파일 없음 / 확인필요', '미대상']:
        st(s.cell(r, 1)); s.cell(r, 1).value = k
        st(s.cell(r, 2), False, 10, 'C00000' if k == '파일 없음 / 확인필요' and cc.get(k) else '000000', None, 'center')
        s.cell(r, 2).value = cc.get(k, 0); r += 1
    st(s.cell(r, 1), True, 10, fill=OK); s.cell(r, 1).value = f'합계 {len(chk)}'
    for col, w in zip('ABC', [30, 12, 10]):
        s.column_dimensions[col].width = w

    # ---- 매핑 테이블 ----
    m = wbo.create_sheet('매핑 테이블')
    m.merge_cells('A1:C1'); m['A1'] = '내용 기반 매핑 (AI가 내용 읽고 생성 · 사람 검토용)'
    st(m['A1'], True, 12, 'FFFFFF', NAVY, 'center')
    for j, h in enumerate(['산출물(매핑 키)', '파일 / URL', '내용 근거']):
        st(m.cell(2, j + 1), True, 10, 'FFFFFF', NAVY, 'center'); m.cell(2, j + 1).value = h
    seen = set(); rr = 3
    for k, v in ev.items():
        f = v[0]
        if f in seen:
            continue
        seen.add(f)
        st(m.cell(rr, 1)); m.cell(rr, 1).value = k
        st(m.cell(rr, 2), fill=BAND if 'http' in f else OK); m.cell(rr, 2).value = f
        st(m.cell(rr, 3), False, 9, '595959'); m.cell(rr, 3).value = v[1] if len(v) > 1 else ''
        rr += 1
    for col, w in zip('ABC', [26, 46, 40]):
        m.column_dimensions[col].width = w
    m.freeze_panes = 'A3'

    # ---- 파일 관점 ----
    d = wbo.create_sheet('파일 관점')
    for j, x in enumerate(['파일 / 폴더', '분류']):
        st(d.cell(1, j + 1), True, 10, 'FFFFFF', NAVY, 'center'); d.cell(1, j + 1).value = x
    cmap = {'미사용': WARN, '잣대/참조(점검대상 아님)': GREY, '사용됨': OK, '사용(추적성)': OK}
    for i, (fn, cat) in enumerate(sorted(fileside, key=lambda x: (x[1] != '미사용', x[0])), start=2):
        st(d.cell(i, 1), False, 9); d.cell(i, 1).value = fn
        st(d.cell(i, 2), cat == '미사용', 9, 'C00000' if cat == '미사용' else '000000', cmap.get(cat))
        d.cell(i, 2).value = cat
    d.column_dimensions['A'].width = 58; d.column_dimensions['B'].width = 24; d.freeze_panes = 'A2'

    # ---- 체크시트 관점 ----
    c = wbo.create_sheet('체크시트 관점')
    for j, x in enumerate(['산출물', '상태', '파일/URL']):
        st(c.cell(1, j + 1), True, 10, 'FFFFFF', NAVY, 'center'); c.cell(1, j + 1).value = x
    order = {'파일 없음 / 확인필요': 0, '시스템(URL)': 1, '파일 연결됨': 2, '미대상': 3}
    cmap2 = {'파일 없음 / 확인필요': WARN, '시스템(URL)': BAND, '파일 연결됨': OK, '미대상': GREY}
    for i, (prod, stt, fu) in enumerate(sorted(chk, key=lambda x: (order.get(x[1], 9), x[0])), start=2):
        st(c.cell(i, 1)); c.cell(i, 1).value = prod
        st(c.cell(i, 2), False, 10, fill=cmap2.get(stt)); c.cell(i, 2).value = stt
        st(c.cell(i, 3), False, 9); c.cell(i, 3).value = fu[:60]
    for col, w in zip('ABC', [28, 18, 50]):
        c.column_dimensions[col].width = w
    c.freeze_panes = 'A2'

    wbo.save(out)
    print(f'매핑 통합 리포트 저장: {out}')
    print('  파일 관점:', dict(fc))
    print('  체크시트 관점:', dict(cc))


if __name__ == '__main__':
    if len(sys.argv) != 5:
        print(__doc__); sys.exit(1)
    main(*sys.argv[1:5])
