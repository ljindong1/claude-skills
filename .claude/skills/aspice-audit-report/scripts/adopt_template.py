#!/usr/bin/env python3
"""새 확정 양식 채택 자동화 — 접수부터 이력 반영까지 한 번에.

새 최신 양식(완성본이어도 됨)을 받으면:
  1) 빈 양식 생성 (답 칸 클리어, 증적 이미지 자동 제거)
  2) Cover 정리 (프로젝트명 비움, 문서이력 데이터 행 클리어)
  3) 이전 보유 양식(assets/latest_template.xlsx)과 문항을 자동 대조해 변경 요약 생성
  4) Template History에 개정 행 자동 추가
     - 첨부본이 이미 더 높은 개정번호를 정식 기재했으면 그대로 채택(자동 추가 안 함)
     - 내용이 바뀌었는데 번호가 그대로면(이력 누락 — JX 실측 사례) 다음 번호로 자동 등재
     - 작성/검토/승인은 공란(공식 승인 후 기입)
  5) assets/latest_template.xlsx 교체 + template_version.json 갱신

usage:
  python adopt_template.py <새양식.xlsx>                    # 스킬 assets에 반영
  python adopt_template.py <새양식.xlsx> --assets <폴더>    # 지정 폴더에 반영(테스트용)
  python adopt_template.py <새양식.xlsx> --dry-run          # 변경 요약만 출력

이후 .skill 재패키징 → 사용자 Save skill 안내는 별도(스킬 폴더는 실행 중 읽기 전용).
"""
import os
import re
import sys
import json
import datetime
import argparse
import tempfile
from copy import copy
import openpyxl
from openpyxl.cell.cell import MergedCell
import config
import template_version as tv
import make_blank_template as mbt


def q_set(wb, kind):
    sh = config.find_sheets(wb)
    out = set()
    if not sh[kind]:
        return out
    ws = wb[sh[kind][0]]
    C = config.detect_cols(ws, kind)
    for r in range(config.DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(r, C["question"]).value
        if v and str(v).strip():
            out.add(" ".join(str(v).split()))
    return out


def prod_set(wb, kind):
    sh = config.find_sheets(wb)
    out = set()
    if not sh[kind]:
        return out
    ws = wb[sh[kind][0]]
    C = config.detect_cols(ws, kind)
    col = C.get("product") or C.get("output_product")
    for r in range(config.DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v and str(v).strip():
            out.add(str(v).strip().splitlines()[0][:40])
    return out


def build_diff_text(old_wb, new_wb):
    lines, n = [], 0
    for kind, label in (("process", "Process Checklist"), ("wp", "WP Checklist")):
        qo, qn = q_set(old_wb, kind), q_set(new_wb, kind)
        add, rem = len(qn - qo), len(qo - qn)
        if add or rem:
            n += 1
            lines.append(f"{n}. {label} 문항 변경 — 추가 {add}건, 삭제/문구변경 {rem}건 (문항 텍스트 기준 자동 대조)")
    po, pn = prod_set(old_wb, "target"), prod_set(new_wb, "target")
    ta, tr = sorted(pn - po), sorted(po - pn)
    if ta or tr:
        n += 1
        detail = []
        if ta:
            detail.append("추가: " + ", ".join(ta[:4]) + ("…" if len(ta) > 4 else ""))
        if tr:
            detail.append("삭제: " + ", ".join(tr[:4]) + ("…" if len(tr) > 4 else ""))
        lines.append(f"{n}. Target 출력 작업 산출물 조정 ({len(po)}→{len(pn)}) — " + " / ".join(detail))
    if not lines:
        lines.append("1. 문항·산출물 변경 없음 (레이아웃/서식 개정)")
    lines.append("※ 이전 보유 양식과의 자동 대조로 생성 — 작성/검토/승인은 공식 승인 후 기입")
    return "\n".join(lines)


def append_history(wb, rev, date, content):
    ws = wb["Template History"]
    # 마지막 개정 데이터 행 탐색
    last = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip().isdigit():
            last = r
    if last is None:
        raise SystemExit("Template History에 기존 개정 행이 없습니다")
    new_r = last + 1
    for c in range(1, 10):
        cell = ws.cell(new_r, c)
        src = ws.cell(last, c)
        cell.font = copy(src.font); cell.border = copy(src.border)
        cell.alignment = copy(src.alignment); cell.fill = copy(src.fill)
        if not isinstance(cell, MergedCell):
            cell.value = None
    for c, v in ((1, rev), (2, date), (3, content)):
        cell = ws.cell(new_r, c)
        if isinstance(cell, MergedCell):
            # 병합 범위의 좌상단이 아니면 해당 병합을 풀고 기록
            for m in list(ws.merged_cells.ranges):
                if (new_r, c) in [(rr, cc) for rr in range(m.min_row, m.max_row + 1)
                                  for cc in range(m.min_col, m.max_col + 1)]:
                    ws.unmerge_cells(str(m))
                    break
            cell = ws.cell(new_r, c)
        cell.value = v
    if not any(str(m) == f"C{new_r}:F{new_r}" for m in ws.merged_cells.ranges):
        try:
            ws.merge_cells(f"C{new_r}:F{new_r}")
        except Exception:
            pass
    ws.row_dimensions[new_r].height = 15 * (content.count("\n") + 2)


def clean_cover(wb):
    if "Cover" not in wb.sheetnames:
        return
    ws = wb["Cover"]
    in_hist = False
    for r in range(1, ws.max_row + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            t = str(v).strip()
            if t.startswith("프로젝트 명"):
                ws.cell(r, c).value = "프로젝트 명 : "
            if t == "버전번호":
                in_hist = True
                hist_header = r
        if in_hist and r > hist_header:
            # 문서이력 데이터 행(V0.1 등)을 비운다
            if any(re.match(r"^V\d", str(ws.cell(r, c).value or "").strip()) for c in range(1, 5)):
                for c in range(1, (ws.max_column or 0) + 1):
                    if ws.cell(r, c).value is not None:
                        ws.cell(r, c).value = None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", help="새 확정 양식(완성본 가능)")
    ap.add_argument("--assets", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets"))
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    assets = os.path.abspath(a.assets)
    cur_tpl = os.path.join(assets, "latest_template.xlsx")
    cur_json = os.path.join(assets, "template_version.json")
    if not os.path.exists(cur_tpl):
        sys.exit(f"보유 양식 없음: {cur_tpl}")
    stored = json.load(open(cur_json, encoding="utf-8")) if os.path.exists(cur_json) else {}

    # 1) 빈 양식 생성(전체 클리어 → 증적 이미지 자동 제거)
    tmp = tempfile.mktemp(suffix=".xlsx")
    mbt.main(a.src, tmp)

    new_wb = openpyxl.load_workbook(tmp)
    old_wb = openpyxl.load_workbook(cur_tpl, data_only=True)

    # 2) Cover 정리
    clean_cover(new_wb)

    # 3) 변경 자동 대조
    diff_text = build_diff_text(old_wb, new_wb)
    print("\n=== 자동 대조 변경 요약 ===")
    print(diff_text)

    # 4) 개정 이력 처리
    in_rev, _ = tv.read_revision(new_wb)
    st_rev = stored.get("revision")
    today = datetime.date.today().strftime("%y.%m.%d")
    if in_rev is not None and st_rev is not None and in_rev > st_rev:
        print(f"\n첨부본이 공식 개정 {in_rev}을 이미 기재 — 이력 자동 추가 없이 그대로 채택")
        new_rev = in_rev
    else:
        new_rev = max(x for x in (in_rev, st_rev, 0) if x is not None) + 1
        print(f"\n내용 변경 대비 이력 미기재 → 개정번호 {new_rev}로 자동 등재 ({today})")
        append_history(new_wb, new_rev, today, diff_text)

    if a.dry_run:
        print("[dry-run] 파일 반영 생략")
        return

    # 5) 반영
    new_wb.save(cur_tpl)
    wb2 = openpyxl.load_workbook(cur_tpl, data_only=True)
    rev, date = tv.read_revision(wb2)
    fp, layout, content = tv.fingerprint(wb2)
    json.dump({
        "revision": rev, "revised": date, "fingerprint": fp, "content": content,
        "source": f"{os.path.basename(a.src)} → adopt_template.py 자동 채택({datetime.date.today()})",
        "stored": str(datetime.date.today()),
        "note": "양식 갱신 절차: 새 확정 양식 접수 → python scripts/adopt_template.py <새양식> → .skill 재패키징 후 Save skill",
    }, open(cur_json, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"반영 완료: {cur_tpl} (개정번호 {rev}) + template_version.json")
    print("다음 단계: .skill 재패키징 후 사용자에게 Save skill 안내")


if __name__ == "__main__":
    main()
