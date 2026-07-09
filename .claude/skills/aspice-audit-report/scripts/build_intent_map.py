#!/usr/bin/env python3
"""체크리스트(.xlsx)에서 문항을 추출해 intent_map 초안을 생성한다 — Phase B.

결정론적 스크립트는 문항을 '뽑아 빈 칸을 만드는' 데까지만 한다.
intent/focus/depth/primary_location/alt_evidence 는 AI가 문항을 읽고 채운다
(references/comment_style.md §10~11). 이렇게 초안→검수 방식으로 지식베이스를
프로젝트에 걸쳐 축적한다(common_mapping_rules.json 과 동일 철학).

usage:
  python build_intent_map.py <체크리스트.xlsx> --out intent_map.json
  python build_intent_map.py <체크리스트.xlsx> --out intent_map.json --merge 기존.json
  python build_intent_map.py <체크리스트.xlsx> --out intent_map.json --sheet "WP Checklist"

--merge: 기존 intent_map의 채워진 항목을 (sheet,no,product) 키로 보존하고,
         새 문항만 빈 칸으로 추가한다(양식 개정 시 재사용).
"""
import argparse
import json
import sys

import openpyxl

import config
import intent_lib


EMPTY_FIELDS = {
    "intent": "",
    "pam_ref": "",
    "focus": "",
    "depth": "",
    "primary_location": "",
    "alt_evidence": [],
}


def _first_line(v):
    return str(v or "").strip().splitlines()[0][:40] if v is not None else ""


def extract(path, only=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = config.find_sheets(wb)
    entries = []
    for kind in ("wp", "process"):
        for name in sh[kind]:
            if only and name not in only:
                continue
            ws = wb[name]
            C = config.detect_cols(ws, kind)
            gcol = C["product"] if kind == "wp" else C["output_product"]
            qcol = C["question"]
            cur = None
            for r in range(config.DATA_START_ROW, ws.max_row + 1):
                no = ws.cell(r, 1).value
                g = ws.cell(r, gcol).value
                if g and str(g).strip():
                    cur = _first_line(g)
                if no is None:
                    continue
                q = ws.cell(r, qcol).value
                e = {"sheet": kind, "no": no, "product": cur or "",
                     "question_excerpt": str(q or "").strip()[:160]}
                e.update({k: (list(v) if isinstance(v, list) else v)
                          for k, v in EMPTY_FIELDS.items()})
                entries.append(e)
    return entries


def merge(new_entries, old_path):
    old = intent_lib.IntentMap.load(old_path)
    kept = 0
    for e in new_entries:
        prev = old.get_exact(e["sheet"], e["no"], e["product"])
        if prev and str(prev.get("intent", "")).strip():
            for k in EMPTY_FIELDS:
                if prev.get(k):
                    e[k] = prev[k]
            if prev.get("pam_ref"):
                e["pam_ref"] = prev["pam_ref"]
            kept += 1
    return kept


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("checklist")
    ap.add_argument("--out", required=True)
    ap.add_argument("--merge", default=None, help="기존 intent_map.json (채워진 항목 보존)")
    ap.add_argument("--sheet", action="append", help="이 시트만 추출")
    a = ap.parse_args()

    entries = extract(a.checklist, a.sheet)
    kept = merge(entries, a.merge) if a.merge else 0
    filled = sum(1 for e in entries if str(e.get("intent", "")).strip())

    doc = {
        "_meta": {
            "source": a.checklist,
            "total": len(entries),
            "filled": filled,
            "note": "intent/focus/depth/primary_location/alt_evidence 는 AI가 문항을 읽고 채운다 (comment_style §10~11).",
        },
        "entries": entries,
    }
    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    print(f"추출 {len(entries)}문항 → {a.out}")
    if a.merge:
        print(f"기존에서 보존(채워짐) {kept}건")
    print(f"채워진 항목 {filled} / 미작성 {len(entries) - filled} "
          f"(미작성은 AI가 intent 등을 채워야 함)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
