#!/usr/bin/env python3
"""AI가 작성한 사람급 코멘트(comments.json)를 결과 체크리스트에 주입한다.

comments.json 형식:
[ {"sheet":"wp","product":"프로젝트 계획서","no":2,
   "comment":"260624 (AI·본문정독): ...","result":"Pass"}, ... ]
- sheet : "wp" 또는 "process"
- sheet_name : (선택) 같은 종류 시트가 여러 개일 때 대상 시트명을 정확히 지정.
               생략 시 --sheet 제한 안에서(없으면 그 종류 전체에서) no/product로 찾는다.
- product: WP의 No는 산출물 간 중복되므로 product 부분문자열로 행을 구분(필수 권장)
- no : 해당 시트의 No 열 값
- result: 선택(Pass/Fail/Conditional pass). 생략 시 Result는 건드리지 않음.

usage:
  python write_comments.py <결과.xlsx> <comments.json>
  python write_comments.py <결과.xlsx> <comments.json> --sheet "Process Checklist_테스트 단계말"
"""
import re
import sys
import json
import argparse
import openpyxl
import config

PASS_TAG_RE = re.compile(r"판정:\s*Pass(?!\s*후보)")


def tag_to_result(comment):
    """AI 1차 의견 직접 기록 정책: 코멘트의 판정 태그를 Result 값으로 매핑."""
    t = str(comment or "")
    if PASS_TAG_RE.search(t) or "Pass 후보" in t:
        return "Pass"
    if "Fail 후보" in t:
        return "Fail"
    if "확인 필요" in t or "확인필요" in t:
        return "확인 필요"
    return None


def main(result, cj, only_sheets=None):
    items = json.load(open(cj, encoding="utf-8"))
    wb = openpyxl.load_workbook(result)
    sh = config.find_sheets(wb)
    n = 0
    missed = []
    for it in items:
        kind = it["sheet"]
        names = sh.get(kind, [])
        if it.get("sheet_name"):
            names = [x for x in names if x == it["sheet_name"]]
        elif only_sheets:
            names = [x for x in names if x in only_sheets]
        if not names:
            missed.append((it.get("product"), it["no"], "대상 시트 없음"))
            continue
        prod = it.get("product")
        hit = False
        for name in names:
            ws = wb[name]
            C = config.detect_cols(ws, kind)
            pcol = C.get("product") if kind == "wp" else None
            for row in range(config.DATA_START_ROW, ws.max_row + 1):
                if ws.cell(row, 1).value != it["no"]:
                    continue
                if prod and pcol and prod not in str(ws.cell(row, pcol).value or ""):
                    continue
                ws.cell(row, C["comments"]).value = it["comment"]
                # 태그-Result 자동 동기화: 태그를 Result 값으로 직접 기록(사람이 확정·수정)
                if not it.get("result"):
                    tr = tag_to_result(it["comment"])
                    if tr:
                        it = dict(it, result=tr)
                if it.get("result"):
                    rv = str(it["result"]).strip()
                    if rv not in config.ALLOWED_RESULTS:
                        missed.append((prod, it["no"], f"허용 외 판정값 {rv!r}"))
                        continue
                    ws.cell(row, C["result"]).value = rv
                    if C.get("applicable"):
                        ws.cell(row, C["applicable"]).value = "Yes"
                n += 1
                hit = True
        if not hit:
            missed.append((it.get("product"), it["no"], "행 못 찾음"))
    wb.save(result)
    print(f"주입된 셀: {n} | 미주입: {len(missed)}")
    for m in missed:
        print("  miss", m)
    return missed


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("result")
    ap.add_argument("comments")
    ap.add_argument("--sheet", action="append", help="이 시트에만 주입(증분 모드)")
    a = ap.parse_args()
    main(a.result, a.comments, a.sheet)
