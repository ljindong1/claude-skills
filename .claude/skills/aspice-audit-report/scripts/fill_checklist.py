#!/usr/bin/env python3
"""빈 양식에 산출물 매핑 + 1차 자동 판정을 채운다.

usage:
  python fill_checklist.py <빈양식.xlsx> <결과.xlsx> --map deliverable_map.json
  # 증분 모드 — 지정 시트만 채운다(다른 시트의 기존 결과 보존):
  python fill_checklist.py <빈양식.xlsx> <결과.xlsx> --map m.json --sheet "Process Checklist_테스트 단계말"

시트는 "있는 것만" 채운다 — Target/WP 없는 Process 전용 양식도 정상 입력.
"""
import re
import json
import argparse
import datetime
import openpyxl
import config
import classification_rules as rules


def norm(s):
    return re.sub(r"[\s()◆\[\]]", "", str(s or "")).lower()


def load_map(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def find_ev(pn, evidence):
    # 가장 구체적인(긴) 키가 이긴다 — 일반 키('회의록')가 특정 키('검토회의록')를 가로채는 것 방지
    best, best_len = None, -1
    for kw, val in evidence.items():
        nk = norm(kw)
        if nk and nk in pn and len(nk) > best_len:
            best, best_len = val, len(nk)
    if best:
        return best[0], (best[1] if len(best) > 1 else "")
    return None


def is_na(product, extra):
    if rules.is_design_phase_target(product):   # 계획서·정적검증은 미대상에서 제외(점검대상 유지)
        return False
    if rules.is_na_product(product):
        return True
    pn = norm(product)
    return any(norm(x) in pn for x in extra)


def pick(sheets, only):
    """only(사용자 지정 시트명들)가 있으면 교집합, 없으면 전부."""
    if not only:
        return sheets
    return [n for n in sheets if n in only]


def fill(blank, out, mapping, only_sheets=None):
    today = datetime.date.today().strftime("%y%m%d")
    ev = mapping.get("evidence", {})
    na_reason = mapping.get("na_default_reason", "단계 미도래/해당없음 점검 범위")
    na_extra = mapping.get("na_products", [])
    wb = openpyxl.load_workbook(blank)
    sh = config.find_sheets(wb)
    if only_sheets:
        known = [n for names in sh.values() for n in names]
        bad = [n for n in only_sheets if n not in known]
        if bad:
            raise SystemExit(f"--sheet 시트 없음/비인식: {bad} (인식된 시트: {known})")
    counts = dict(target=0, wp_no=0, wp_pass=0, wp_chk=0, wp_unk=0, pc_no=0, pc_chk=0, pc_cond=0)
    filled_sheets = []

    # Target (있을 때만)
    for name in pick(sh["target"], only_sheets):
        ws = wb[name]
        T = config.detect_cols(ws, "target")
        filled_sheets.append(name)
        for row in range(config.DATA_START_ROW, ws.max_row + 1):
            prod = ws.cell(row, T["product"]).value
            if not prod:
                continue
            hit = find_ev(norm(prod), ev)
            if hit:
                ws.cell(row, T["file"]).value = hit[0]
                ws.cell(row, T["bigo"]).value = "[점검 대상]"
            elif is_na(prod, na_extra):
                ws.cell(row, T["bigo"]).value = "[점검 미대상] " + na_reason
            else:
                ws.cell(row, T["bigo"]).value = "[점검 대상] 확인 필요 - 자동 식별 안됨"
            counts["target"] += 1

    # WP (있을 때만)
    for name in pick(sh["wp"], only_sheets):
        ws = wb[name]
        W = config.detect_cols(ws, "wp")
        apc = W.get("applicable")
        if not apc:
            print(f"  [스킵] {name!r}: 구버전 레이아웃(Applicable 없음) — 신규 작성 금지, 읽기 전용 보존")
            continue
        filled_sheets.append(name)
        for row in range(config.DATA_START_ROW, ws.max_row + 1):
            no = ws.cell(row, 1).value
            ques = ws.cell(row, W["question"]).value
            if no is None or not ques:
                continue
            prod = ws.cell(row, W["product"]).value
            hit = find_ev(norm(prod), ev)
            if hit:
                if apc:
                    ws.cell(row, apc).value = "Yes"
                evstr = (hit[0] + " (" + hit[1] + ")") if hit[1] else hit[0]
                res, com = rules.judge(ques, evstr)
                com = com.replace("(AI 1차)", today + " (AI 1차)")
                if res:
                    ws.cell(row, W["result"]).value = res
                    counts["wp_pass"] += 1
                else:
                    ws.cell(row, W["result"]).value = "확인 필요"
                    counts["wp_chk"] += 1
                ws.cell(row, W["comments"]).value = com
            elif is_na(prod, na_extra):
                if apc:
                    ws.cell(row, apc).value = "No"
                pname = str(prod or "").strip().splitlines()[0][:30]
                note = ""
                ex = find_ev(norm(prod), ev)
                if ex:
                    note = " ※ 산출물 '" + ex[0] + "' 폴더 존재 — 범위 제외가 맞는지 확인 필요"
                ws.cell(row, W["comments"]).value = "[점검 미대상] " + pname + ": " + na_reason + note
                counts["wp_no"] += 1
            else:
                # 매핑된 파일이 없어도 미대상이 아니면 점검대상이다 → Applicable=Yes, 판정은 보류(확인 필요)
                if apc:
                    ws.cell(row, apc).value = "Yes"
                ws.cell(row, W["result"]).value = "확인 필요"
                ws.cell(row, W["comments"]).value = today + " (AI 1차): 대상 산출물 자동 식별 안됨 - 확인 필요(폴더 외/codebeamer 가능)"
                counts["wp_unk"] += 1

    # Process — Yes/No 스코핑: 실행시점 후행단계/미대상 → No, 출력 산출물 식별 → Yes, 미식별 → 비움(확인 필요)
    for name in pick(sh["process"], only_sheets):
        ws = wb[name]
        P = config.detect_cols(ws, "process")
        apc = P.get("applicable")
        if not apc:
            print(f"  [스킵] {name!r}: 구버전 레이아웃(Applicable 없음) — 신규 작성 금지, 읽기 전용 보존")
            continue
        filled_sheets.append(name)
        for row in range(config.DATA_START_ROW, ws.max_row + 1):
            ques = ws.cell(row, P["question"]).value
            if not ques:
                continue
            out_p = ws.cell(row, P["output_product"]).value
            timing = ws.cell(row, P["timing"]).value
            if rules.is_later_phase_timing(timing):
                t18 = str(timing or "")[:18]
                if apc:
                    ws.cell(row, apc).value = "No"
                ws.cell(row, P["comments"]).value = "[점검 미대상] 실행시점 후행단계(" + t18 + ") - 단계 미도래"
                counts["pc_no"] += 1
            elif is_na(out_p, na_extra):
                if apc:
                    ws.cell(row, apc).value = "No"
                pname = str(out_p or "").strip().splitlines()[0][:30]
                note = ""
                ex = find_ev(norm(out_p), ev)
                if ex:
                    note = " ※ 산출물 '" + ex[0] + "' 폴더 존재 — 범위 제외가 맞는지 확인 필요"
                ws.cell(row, P["comments"]).value = "[점검 미대상] " + pname + ": " + na_reason + note
                counts["pc_no"] += 1
            else:
                hit = find_ev(norm(out_p), ev)
                if (rules.is_conditional_activity(timing) or rules.is_event_driven(out_p, ques)) and not hit:
                    # 조건부 활동(주기 도래/발생 시) + 폴더에 수행 흔적 없음 → 스코핑 보류(사람 확정)
                    why = ("실행시점 '" + str(timing or "").strip()[:20] + "'") if rules.is_conditional_activity(timing) else "사건발생형 활동(형상감사/CCB/변경 등)"
                    ws.cell(row, P["comments"]).value = (today + " (AI 1차): [조건부 활동] " + why
                        + " — 폴더에 수행 흔적 없음. 발생/도래 여부 확인 필요(미발생이면 No, 발생했으면 Yes 후 점검)")
                    counts["pc_cond"] += 1
                elif hit:
                    if apc:
                        ws.cell(row, apc).value = "Yes"
                    ws.cell(row, P["result"]).value = "Pass"
                    ws.cell(row, P["comments"]).value = (today + " (AI 1차): Pass 후보 — 출력 산출물 식별("
                        + hit[0] + "). 승인·등록 등 시스템 수행 기록만 확인 필요")
                    counts["pc_chk"] += 1
                else:
                    if apc:
                        ws.cell(row, apc).value = "Yes"
                    ws.cell(row, P["result"]).value = "확인 필요"
                    ws.cell(row, P["comments"]).value = (today
                        + " (AI 1차): 확인 필요 — 출력 산출물 미식별. 절차 수행 근거(PMS 승인/시스템 기록) 직접 확인 필요")
                    counts["pc_chk"] += 1

    if not filled_sheets:
        raise SystemExit("채울 시트가 없습니다 (--sheet 지정 또는 양식 시트명 확인)")

    add_notice(wb, counts, today, filled_sheets)
    wb.save(out)
    issues = integrity(wb, sh, only_sheets)
    print("저장:", out)
    print("채운 시트:", filled_sheets)
    print("집계:", counts)
    print("정합성 이슈:", len(issues))
    for i in issues[:10]:
        print("  ", i)


def integrity(wb, sh, only_sheets=None):
    """미대상인데 결과 존재, 코멘트 누락, 허용 외 판정값 등 모순 검사."""
    issues = []
    for kind in ("wp", "process"):
        for name in pick(sh[kind], only_sheets):
            ws = wb[name]
            C = config.detect_cols(ws, kind)
            apc = C.get("applicable")
            for row in range(config.DATA_START_ROW, ws.max_row + 1):
                no = ws.cell(row, 1).value
                ques = ws.cell(row, C["question"]).value
                if no is None or not ques:
                    continue
                ap = ws.cell(row, apc).value if apc else None
                rs = ws.cell(row, C["result"]).value
                cm = ws.cell(row, C["comments"]).value
                if ap == "No" and rs:
                    issues.append((name, row, "No인데 Result 존재"))
                if rs and str(rs).strip() not in config.ALLOWED_RESULTS:
                    issues.append((name, row, f"허용 외 판정값: {rs!r}"))
                if "판정: Pass" in str(cm or "") and not (rs and str(rs).strip()):
                    issues.append((name, row, "'판정: Pass' 태그인데 Result 빈칸 — 동기화 필요"))
                if not cm:
                    issues.append((name, row, "코멘트 없음"))
    return issues


def add_notice(wb, counts, today, filled_sheets):
    from openpyxl.styles import Font, PatternFill, Alignment
    name = "AI점검_안내"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 92
    ws.sheet_view.showGridLines = False
    chk = counts["wp_chk"] + counts["wp_unk"]
    lines = [
        ("AI 1차 자동 점검 결과 — 안내", True, 14, "1F3864", "FFFFFF"),
        (today + " · 본 결과는 AI 1차 자동 점검이며 최종 판정은 사람이 검증해야 함", True, 10, "000000", "FFF2CC"),
        ("· 이번에 채운 시트: " + ", ".join(filled_sheets) + " (그 외 시트의 기존 결과는 보존됨)", False, 10, "000000", None),
        ("· Review Result의 Pass/Fail/확인 필요는 AI 1차 의견이며 최종 확정은 사람이 한다(코멘트의 판정 태그로 근거 수준 구분)", True, 10, "000000", "FFF2CC"),
        ("· Pass = 문서에 해당 절·항목이 존재함(형식·존재성) 수준이며 내용 적정성은 미검증", False, 10, "000000", None),
        ("· Conditional pass(조건부 합격)는 합격 계열로 취급하되 조건 이행 여부는 사람이 확인", False, 10, "000000", None),
        ("· 승인·배포·추적성·일관성 등은 자동 Pass 금지 → 확인 필요로 보류", False, 10, "000000", None),
        ("· 사내 시스템 URL은 접근 불가 → 확인 필요", False, 10, "000000", None),
        ("· Process Yes/No는 출력 산출물 식별·실행시점 기반 1차 스코핑 (사람 확정 필요)", False, 10, "000000", None),
    ]
    if counts["wp_pass"] + counts["wp_no"] + chk:
        lines.append(("· 요약 WP: Pass " + str(counts["wp_pass"]) + " / 미대상 " + str(counts["wp_no"]) + " / 확인필요 " + str(chk), False, 10, "000000", None))
    if counts["pc_no"] + counts["pc_chk"] + counts["pc_cond"]:
        lines.append(("· 요약 Process: 미대상 " + str(counts["pc_no"]) + " / Pass후보·확인필요 " + str(counts["pc_chk"]) + " / 조건부(발생여부 확인) " + str(counts["pc_cond"]), False, 10, "000000", None))
    for i, (t, b, sz, col, fill) in enumerate(lines, start=2):
        c = ws.cell(i, 2, t)
        c.font = Font(name="맑은 고딕", bold=b, size=sz, color=col)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        ws.row_dimensions[i].height = 20


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("blank")
    ap.add_argument("out")
    ap.add_argument("--map", required=True)
    ap.add_argument("--sheet", action="append", help="이 시트만 채운다(증분 모드, 반복 지정 가능)")
    a = ap.parse_args()
    fill(a.blank, a.out, load_map(a.map), a.sheet)
