#!/usr/bin/env python3
"""오염 검사: AI 코멘트가 사람 코멘트를 근거 없이 복제(참조 오염)했는지 검증 단계에서 검사.

배경: 학습 검증에서 AI가 폴더 산출물로 알 수 없는 사실(PMS 승인, P2 변경, Finding 유무 등)을
사람 코멘트에서 가져와 자기 근거처럼 쓴 사례가 실측됨(CN8 No.91·28·14). 이 스크립트는
사람 작성본과 AI 결과를 행 단위로 대조해 오염 의심 항목을 플래그한다.

usage:
  python check_contamination.py <AI결과.xlsx> <사람작성본.xlsx> [--sheet 시트명 ...]

판정:
  [HIGH]  텍스트 유사도 >= 0.55 — 사람 코멘트를 사실상 복제
  [FACT]  시스템 사실 키워드(PMS 승인·그룹웨어·P2 변경·Finding 등)가 사람·AI 양쪽에 있고
          유사도 >= 0.25 — 폴더로 알 수 없는 사실의 전이 의심
종료코드: 플래그 있으면 1, 없으면 0. 블라인드 점검이면 사람 작성본이 없으므로 실행 대상 아님.
"""
import re
import sys
import argparse
import difflib
import openpyxl
import config

SYSTEM_FACTS = ["PMS 승인", "PMS WBS", "PMS 등록", "그룹웨어", "업무연락", "P2 변경",
                "Finding 없", "파인딩 없", "레드마인", "결재", "품의", "전산 등록"]


ASSERT_RE = re.compile(r"(확인됨|확인함|확인되|승인됨|승인 받|득함|완료됨|등록됨|존재함|이행 중)")


def asserts_fact(a, k):
    # 시스템 사실 키워드가 '단정 문맥'으로 쓰였는지 (단순 "접근 필요" 언급은 오염 아님)
    i = a.find(k)
    while i != -1:
        if ASSERT_RE.search(a[i:i + len(k) + 16]):
            return True
        i = a.find(k, i + 1)
    return False


def norm(c):
    t = str(c or "")
    t = re.sub(r"^\d{6}\s*\([^)]*\)\s*:?", "", t.strip())   # 날짜·마커 제거
    t = re.sub(r"^\d{4}-\d{2}-\d{2}\s*:?", "", t.strip())
    t = re.sub(r"\s+", " ", t)
    return t


def rows(ws, kind):
    C = config.detect_cols(ws, kind)
    gcol = C["product"] if kind == "wp" else C["output_product"]
    cur = None
    for r in range(config.DATA_START_ROW, ws.max_row + 1):
        no = ws.cell(r, 1).value
        g = ws.cell(r, gcol).value
        if g and str(g).strip():
            cur = str(g).strip().splitlines()[0][:24]
        if no is None:
            continue
        yield no, cur, ws.cell(r, C["comments"]).value


def main(ai, human, only=None):
    A = openpyxl.load_workbook(ai, data_only=True)
    H = openpyxl.load_workbook(human, data_only=True)
    sh = config.find_sheets(A)
    flags = []
    for kind in ("wp", "process"):
        for name in sh[kind]:
            if only and name not in only:
                continue
            if name not in H.sheetnames:
                continue
            hmap = {}
            for no, g, c in rows(H[name], kind):
                hmap[(g, no) if kind == "wp" else no] = norm(c)
            for no, g, c in rows(A[name], kind):
                key = (g, no) if kind == "wp" else no
                a, h = norm(c), hmap.get(key, "")
                if not a or not h:
                    continue
                ratio = difflib.SequenceMatcher(None, a, h).ratio()
                facts = [k for k in SYSTEM_FACTS if k in a and k in h and asserts_fact(a, k)]
                if any(k in a for k in ("사람 의견", "사람 판정", "사람 코멘트")):
                    flags.append(("REF", name, no, g, round(ratio, 2), ["사람 의견 명시 참조"]))
                elif ratio >= 0.55:
                    flags.append(("HIGH", name, no, g, round(ratio, 2), facts))
                elif facts and ratio >= 0.25:
                    flags.append(("FACT", name, no, g, round(ratio, 2), facts))
    if flags:
        print(f"[오염 의심] {len(flags)}건 — AI 코멘트가 사람 코멘트를 근거 없이 복제했는지 확인하라")
        for lv, name, no, g, ratio, facts in flags[:40]:
            print(f"  [{lv}] {name} No.{no} ({str(g)[:20]}) 유사도={ratio}" + (f" 시스템사실={facts}" if facts else ""))
        print("→ 해당 항목은 폴더 산출물 근거로 재작성하거나 '확인 필요'로 정직하게 남겨라.")
        return 1
    print("[통과] 오염 의심 항목 없음 — AI 코멘트가 사람 코멘트와 독립적임")
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("ai")
    ap.add_argument("human")
    ap.add_argument("--sheet", action="append")
    a = ap.parse_args()
    sys.exit(main(a.ai, a.human, a.sheet))
