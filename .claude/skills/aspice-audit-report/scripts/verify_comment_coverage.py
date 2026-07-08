#!/usr/bin/env python3
"""사람급 코멘트 전수 커버리지 최종 확인 (양식에 있는 WP·Process 시트 전부).

목적: 리포트의 점검대상(Applicable=Yes) 항목이 '하나도 빠짐없이' 사람급(AI 심층)
코멘트를 가졌는지 검사하고, 빠진/얕은(1차 템플릿) 항목을 시트·산출물별로 보고한다.
사람 작성본이 있으면 사람이 코멘트한 항목 기준으로도 대조(같은 이름 시트끼리).

usage:
  python verify_comment_coverage.py <결과.xlsx> [사람작성본.xlsx] [--sheet 시트명 ...]

- WP 시트가 없는 Process 전용 양식이면 Process만 검사한다.
- --sheet 지정 시 그 시트만 검사(증분 모드 — 기존 차수 시트는 제외).
- Applicable 열이 없는 구버전 시트는 Result 존재 여부로 Yes를 추정한다.
종료코드: 미달(빠짐 또는 비심층)이 있으면 1, 전부 충족이면 0.
'심층' 판정은 코멘트에 AI 정독 마커(DEEP_MARKERS)가 있는지로 본다.
"""
import re
import sys
import datetime
import argparse
import openpyxl
import config
import classification_rules as rules
from collections import defaultdict

DEEP_MARKERS = ["AI·본문", "AI·결과서", "AI·검토", "AI·정독", "AI·구조", "AI·추적",
                "AI·데이터", "AI·확인", "본문+검토", "본문정독", "검토결과서"]

# 근거 인용으로 인정하는 패턴: 파일명 확장자 / 절·표·시트 번호 / 버전·개정 / 정직한 한계 표기
EVIDENCE_RE = re.compile(
    r"\.(docx|xlsx|xls|pdf|md)|절\s?[\'\d]|표\s?\d|시트|Rev|V\d|개정|이력|"
    r"문서 부재|폴더 부재|폴더 외|확인 필요|확인필요|미대상")
DATE_RE = re.compile(r"^(\d{6})\s*\(")
# Pass 판정인데 코멘트가 요구 산출물의 부재를 말하면 모순 (실측: No.32 위험오판)
ABSENT_RE = re.compile(r"부재|작성되지 않|작성·승인·배포되지 않|미작성|승인되지 않|배포되지 않|확인 불가|확인되지 않|존재하지 않")
# AI가 '부재를 확인'한 문구 — 이런 코멘트에는 Fail 후보 태그가 있어야 한다(JX 실측: 부재 인지 후 확인필요로만 남김)
VERIFIED_ABSENT_RE = re.compile(r"확인되지 않|이력 없|이력이 없|기록 없음|기록이 없|부재하")
# 문항 핵심(대사)을 미확인이라 쓰면서 Pass를 다는 자기모순 (JX 실측: 사람 Fail 6건 중 4건)
UNVERIFIED_CORE_RE = re.compile(r"(정합성|일관성|완전성|추적성|추적)[은는이가의도\s·]{0,6}[^.。]{0,20}(확인 필요|확인필요|확인 권고|권고|미확인|확인 불가)")
PASS_TAG_RE = re.compile(r"판정:\s*Pass(?!\s*후보)")


def expected_result(ct):
    """AI 1차 의견 직접 기록 정책의 태그→Result 기대값."""
    if PASS_TAG_RE.search(ct) or "Pass 후보" in ct:
        return "Pass"
    if "Fail 후보" in ct:
        return "Fail"
    if "확인 필요" in ct or "확인필요" in ct:
        return "확인 필요"
    return None

# Process 점검대상 코멘트에 필수인 판정 제안 태그 (없으면 비심층 처리)
JUDGE_TAGS = ["판정: Pass", "Pass 후보", "Fail 후보", "확인 필요", "확인필요", "[조건부 활동]", "미대상"]


def has_judgement(c):
    return any(k in str(c or "") for k in JUDGE_TAGS)


def is_deep(c):
    # 마커 + 근거 인용이 모두 있어야 심층으로 인정 (마커만으로는 불인정)
    t = str(c or "")
    return bool(t) and any(m in t for m in DEEP_MARKERS) and bool(EVIDENCE_RE.search(t))


def date_issue(c):
    # 코멘트 머리의 YYMMDD 마커 연도가 올해가 아니면 오기(연도 오타)로 플래그
    m = DATE_RE.match(str(c or "").strip())
    if not m:
        return None
    if m.group(1)[:2] != datetime.date.today().strftime("%y"):
        return m.group(1)
    return None


def rows(ws, kind):
    C = config.detect_cols(ws, kind)
    gcol = C["product"] if kind == "wp" else C["output_product"]
    apc, rc = C.get("applicable"), C["result"]
    cur = None
    for r in range(config.DATA_START_ROW, ws.max_row + 1):
        no = ws.cell(r, 1).value
        g = ws.cell(r, gcol).value
        if g and str(g).strip():
            cur = str(g).strip().splitlines()[0][:24]
        if no is None:
            continue
        if apc:
            ap = str(ws.cell(r, apc).value or "").strip()
        else:
            ap = "Yes" if str(ws.cell(r, rc).value or "").strip() else "No"
        yield no, cur, ap, ws.cell(r, C["comments"]).value, str(ws.cell(r, rc).value or "").strip()


def main(result, human=None, only=None):
    wb = openpyxl.load_workbook(result, data_only=True)
    sh = config.find_sheets(wb)
    hwb = openpyxl.load_workbook(human, data_only=True) if human else None
    ok = True
    grand = [0, 0]
    human_gap = []
    checked = 0

    for kind in ("wp", "process"):
        for name in sh[kind]:
            if only and name not in only:
                continue
            ws = wb[name]
            checked += 1
            by = defaultdict(lambda: {"t": 0, "deep": 0, "shallow": [], "miss": []})
            date_bad, scope_warn, sync_bad = [], [], []
            for no, g, ap, cmt, res in rows(ws, kind):
                d = date_issue(cmt)
                if d:
                    date_bad.append((no, d))
                gg = str(g or "")
                if ap == "No":
                    if rules.is_ongoing_activity(gg):
                        scope_warn.append((no, gg[:18], "상시 운영 활동(문제해결/이슈/모니터링)은 전 단계 점검대상 — 미대상 재검토"))
                    elif ("단위" in gg or "정적" in gg) and any(k in str(cmt or "") for k in ("후행", "단계말", "범위 밖", "테스트")):
                        scope_warn.append((no, gg[:18], "단위 검증·정적 분석류는 설계 단계말 점검대상 — 미대상 재검토"))
                    elif "계획" in gg and any(k in str(cmt or "") for k in ("후행", "단계말", "범위 밖", "테스트 실행")):
                        scope_warn.append((no, gg[:18], "계획서류는 설계 단계 점검대상 — 후행 단계 사유로 미대상 처리 금지"))
                if ap != "Yes":
                    if "Fail 후보" in str(cmt or ""):
                        scope_warn.append((no, gg[:18], "[모순] 미대상/조건부(발생 미확인) 항목에 Fail 후보 — 발생 여부를 모르면 Fail 후보 성립 불가"))
                    continue
                exp = expected_result(str(cmt or ""))
                if exp and res != exp:
                    sync_bad.append((no, exp, res or "빈칸"))
                if res in config.PASS_FAMILY and "폴더 외" not in str(cmt or "") and ABSENT_RE.search(str(cmt or "")):
                    scope_warn.append((no, gg[:18], "[모순] Result=Pass인데 코멘트가 산출물 부재를 언급 — 부재 확인이면 Fail이 정답"))
                if (res in config.PASS_FAMILY or exp == "Pass") and UNVERIFIED_CORE_RE.search(str(cmt or "")):
                    scope_warn.append((no, gg[:18], "[모순] 문항 핵심(정합성/일관성/추적 대사)을 미확인이라 쓰면서 Pass — '확인 필요'가 정답 (comment_style §7 금지 조건)"))
                ct = str(cmt or "")
                if ("Fail 후보" not in ct and "폴더 외" not in ct
                        and VERIFIED_ABSENT_RE.search(ct)):
                    scope_warn.append((no, gg[:18], "[부재-태그 모순] 부재/미확인을 확인한 코멘트인데 Fail 후보 태그 없음 — 부재 확인 시 Fail 후보는 의무('폴더 외 문서'로 단정 불가하면 그 사유 명기)"))
                d = by[g]
                d["t"] += 1
                if not (cmt and str(cmt).strip()):
                    d["miss"].append(no)
                elif is_deep(cmt) and (kind != "process" or has_judgement(cmt)):
                    d["deep"] += 1
                else:
                    d["shallow"].append(no)
            st = sum(v["t"] for v in by.values())
            sd = sum(v["deep"] for v in by.values())
            grand[0] += st; grand[1] += sd
            print(f"\n=== [{name}] 점검대상 {st} / 심층 {sd} ({(sd/st*100 if st else 0):.1f}%) ===")
            print(f"{'산출물/그룹':30s} 대상  심층  비심층  빠짐")
            for g in by:
                d = by[g]; bad = len(d["shallow"]) + len(d["miss"])
                if bad:
                    ok = False
                print(f"{str(g)[:28]:30s} {d['t']:4d} {d['deep']:5d} {len(d['shallow']):6d} {len(d['miss']):5d}" + ("  ◀" if bad else ""))
                if d["miss"]:
                    print(f"      빠짐 No: {d['miss']}")
                if d["shallow"]:
                    print(f"      비심층 No: {d['shallow']}")
            if date_bad:
                ok = False
                print(f"  [날짜 오기] 마커 연도가 올해가 아님 {len(date_bad)}건: {date_bad[:10]}")
            if sync_bad:
                ok = False
                print(f"  [태그-Result 불일치] 코멘트 태그와 Result 값이 다름 {len(sync_bad)}건 (No, 기대값, 실제): {sync_bad[:12]} — write_comments 동기화로 해결")
            if scope_warn:
                ok = False
                print(f"  [스코핑 재검토 필요] {len(scope_warn)}건:")
                for no, gg, why in scope_warn[:15]:
                    print(f"      No.{no} ({gg}) — {why}")

            if hwb is not None and name in hwb.sheetnames:
                hws = hwb[name]
                aip = {}
                ai_no = {}
                for no, g, ap, c, _res in rows(ws, kind):
                    aip[(g, no)] = c
                    ai_no[no] = c
                for no, g, ap, cmt, _hres in rows(hws, kind):
                    if ap == "Yes" and cmt and str(cmt).strip():
                        # WP는 (product,no), Process는 no 유일이라 no로
                        v = aip.get((g, no)) if kind == "wp" else ai_no.get(no)
                        if not is_deep(v):
                            human_gap.append((name, g, no))

    if not checked:
        print("검사할 WP/Process 시트가 없습니다 (--sheet/양식 확인)")
        return 2

    print(f"\n>>> 전체 점검대상 {grand[0]} / 심층 {grand[1]} ({(grand[1]/grand[0]*100 if grand[0] else 0):.1f}%)")
    if human:
        print(f">>> 사람 코멘트 대비 AI 비심층/빠짐: {len(human_gap)}건")
        for x in human_gap[:60]:
            print("   ", x)

    if ok and not human_gap:
        print("\n[PASS] 검사한 시트의 점검대상 전 항목이 사람급(심층) 코멘트를 보유함.")
        return 0
    print("\n[FAIL] 보완 필요 — 위 '빠짐/비심층' 항목의 본문·검토결과서를 정독해 코멘트 작성 후 재주입하라.")
    return 1


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("result")
    ap.add_argument("human", nargs="?", default=None)
    ap.add_argument("--sheet", action="append", help="이 시트만 검사(증분 모드)")
    a = ap.parse_args()
    sys.exit(main(a.result, a.human, a.sheet))
