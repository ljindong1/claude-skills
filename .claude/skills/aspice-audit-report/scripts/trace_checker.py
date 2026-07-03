#!/usr/bin/env python3
"""codebeamer Traceability Browser export(.xlsx)에서 미연결(추적 누락) 항목을 센다.

추적성/일관성 점검을 네트워크 없이 폴더 export로 자동화한다.
구조 가정: 소스 ID는 좌측 컬럼[~col1], 상태는 그 뒤, 타깃은 우측(없으면 "--"),
비고(예: '아키텍처에 할당되지 않은 사유 확인 필요')는 우측 끝 텍스트 컬럼.

usage:
  python trace_checker.py <폴더 또는 파일...> [--out trace_evidence.json]
"""
import re
import sys
import glob
import json
import os
import argparse
import openpyxl

TOKEN = re.compile(r'\[([A-Za-z][A-Za-z0-9]*)-(\d+)\]')
STATUSES = {'Accepted', 'Rejected', 'Information', 'In Review', 'Reviewed',
            'Unset', 'Proposed', 'New', 'Resolved', 'Closed', 'Draft'}
ACTIVE = {'Accepted', 'In Review', 'Reviewed', 'New', 'Proposed'}  # 추적돼야 하는 상태


def pair_label(path):
    b = os.path.basename(path)
    b = re.sub(r'\.xlsx?$', '', b)
    b = re.sub(r'_[가-힣]+ ?검토.*$', '', b)   # 검토자 꼬리 제거
    b = re.sub(r'[-_]\d{6}.*$', '', b)         # 날짜 꼬리 제거
    b = re.sub(r'.*SWRC[-_]', '', b)            # 프로젝트 접두 제거
    return b.replace('_', '→').replace('-', '→', 1)


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    sources = {}
    cur = None
    for r in ws.iter_rows(values_only=True):
        cells = ['' if c is None else str(c) for c in r]
        toks = [(i, TOKEN.search(c)) for i, c in enumerate(cells) if TOKEN.search(c)]
        if not toks:
            continue
        leftcol = min(i for i, _ in toks)
        if leftcol <= 3:  # 소스 행
            key = toks[0][1].group(0)
            status = next((c.strip() for c in cells if c.strip() in STATUSES), '')
            tgts = [m.group(0) for i, m in toks if i > leftcol]
            note = next((c.strip() for i, c in enumerate(cells)
                         if i >= 8 and c.strip() and not TOKEN.search(c)
                         and c.strip() not in STATUSES and c.strip() != '--'), '')
            sources[key] = {'status': status, 'traced': bool(tgts), 'note': note}
            cur = key
        elif cur:  # 연속 행(추가 타깃)
            sources[cur]['traced'] = True
    return sources


def summarize(path):
    src = parse_file(path)
    total = len(src)
    untraced = [k for k, v in src.items() if not v['traced']]
    gaps = [k for k in untraced if src[k]['status'] in ACTIVE]  # 활성 상태인데 미연결 = 진짜 갭
    flagged = [k for k in untraced if src[k]['note']]
    return {
        'file': os.path.basename(path), 'pair': pair_label(path),
        'total_sources': total, 'untraced': len(untraced),
        'gaps_active': len(gaps), 'gap_ids': gaps[:50],
        'flagged_note': len(flagged),
        'by_status_untraced': {s: sum(1 for k in untraced if src[k]['status'] == s)
                               for s in sorted({src[k]['status'] for k in untraced})},
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('paths', nargs='+')
    ap.add_argument('--out', default=None)
    a = ap.parse_args()
    files = []
    for p in a.paths:
        if os.path.isdir(p):
            files += glob.glob(os.path.join(p, '*raceability*'))
        else:
            files.append(p)
    results = [summarize(f) for f in sorted(files)]
    for r in results:
        print(f"[{r['pair']}] 소스 {r['total_sources']} / 미연결 {r['untraced']} "
              f"(활성갭 {r['gaps_active']}, 비고플래그 {r['flagged_note']}) "
              f"상태별미연결={r['by_status_untraced']}")
    if a.out:
        with open(a.out, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print('저장:', a.out)


if __name__ == '__main__':
    main()
